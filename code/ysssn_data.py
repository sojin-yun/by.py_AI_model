import numpy as np
import os
import openpyxl
from xml.etree.ElementTree import parse
import logging
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.patches as patches
import random
import shutil
import itk, vtk
from torch.functional import split
# from vtk.util.numpy_support import vtk_to_numpy, numpy_to_vtk

ImageType_SS =  itk.Image[itk.ctype('signed short'), 3] #-32768~32767
ImageType_F = itk.Image[itk.ctype('float'), 3] #3.4E-38(-3.4*10^38) ~ 3.4E+38(3.4*10^38) (7digits)


def get_itkImage_from_dicom(filepath, flag_meta=False, flag_info=False):
    namesGenerator = itk.GDCMSeriesFileNames.New()
    namesGenerator.SetUseSeriesDetails(True)
    namesGenerator.AddSeriesRestriction("0008|0021")
    namesGenerator.SetGlobalWarningDisplay(False)
    namesGenerator.SetDirectory(filepath)
    seriesUID = namesGenerator.GetSeriesUIDs()
    # print('seriesUID: ', seriesUID)

    # There are several serieses, but take one with the largest files.
    max_num_file, max_idx = 0, 0
    for i, uid in enumerate(seriesUID):
        fileNames = namesGenerator.GetFileNames(uid)
        num_file = len(fileNames)
        if num_file >= max_num_file: 
            max_num_file = num_file
            max_idx = i
    
    fileNames = namesGenerator.GetFileNames(seriesUID[max_idx])
    if len(fileNames) <= 0:
        return None
    
    # Load dicom
    reader = itk.ImageSeriesReader[ImageType_SS].New()
    dicomIO = itk.GDCMImageIO.New()
    reader.SetImageIO(dicomIO)
    reader.SetFileNames(fileNames)
    reader.Update()
    itkImage = reader.GetOutput()

    if flag_meta is True: return itkImage, reader
    elif flag_meta is False and flag_info is True: return dicomIO, fileNames, reader.GetMetaDataDictionaryArray()
    else: return itkImage

def judge_axial(fileNames):
    answer = True
    for i, fileName in enumerate(fileNames):
        gdcmIO = itk.GDCMImageIO.New()
        single_reader = itk.ImageFileReader[ImageType_SS].New()
        single_reader.SetImageIO(gdcmIO)
        single_reader.SetFileName(fileName)
        single_reader.Update()

        single_tag = single_reader.GetImageIO().GetMetaDataDictionary()
        try:
            iop_split = single_tag["0020|0037"].split("\\")
            if int(iop_split[4][0]) != 1:
                answer = False
                break
                
            if i>3: break
        except:
            return None
    
    return answer
        
def ThresholdItkImage(itkImage, below, upper):
    thresholder = itk.ThresholdImageFilter[ImageType_SS].New()
    thresholder.SetInput(itkImage)
    thresholder.ThresholdBelow(below)
    thresholder.SetOutsideValue(below)
    thresholder.Update()

    upperThresholder = itk.ThresholdImageFilter[ImageType_SS].New()
    upperThresholder.SetInput(thresholder.GetOutput())
    upperThresholder.ThresholdAbove(upper)
    upperThresholder.SetOutsideValue(upper)  
    upperThresholder.Update()
    itkImage_threshold = upperThresholder.GetOutput()

    return itkImage_threshold

def NormalizeItkImage(itkImage, normalization):
    '''
    itkImage type: input-type of input, output-float
    '''
    rescaler = itk.RescaleIntensityImageFilter[type(itkImage), ImageType_F].New()
    rescaler.SetOutputMinimum(0.0)
    rescaler.SetOutputMaximum(float(normalization))
    rescaler.SetInput(itkImage)
    rescaler.Update()
    itkImage_normalized = rescaler.GetOutput()

    return itkImage_normalized

def TH_IMG_Dicom(dir_dicom, slice_num):
    save_dir = "Z:/Backup/Users/kys/BoundingBox/data/TH_Img/"
    itkImage = get_itkImage_from_dicom(dir_dicom)

    itkImage_threshold = ThresholdItkImage(itkImage, below=-150, upper=250)

    itkImage_normalized = NormalizeItkImage(itkImage_threshold, 1.0)

    arr_ct_procesed = itk.GetArrayFromImage(itkImage_normalized)

    plt.imshow(arr_ct_procesed[slice_num], cmap = 'gray')
    plt.tight_layout()
    plt.margins(0,0)
    plt.axis("Off")  
    plt.savefig(f"{save_dir}{dir_dicom.split('/')[8]}_{dir_dicom.split('/')[9][-3:]}_{slice_num}.png", bbox_inches='tight')


def prepare_data(dir_data, dir_result, sigma):
    NUM_Landmark = 2
    if os.path.isdir(dir_result) is False: os.mkdir(dir_result)

    # log
    path_log = os.path.join(dir_result, 'log.txt')
    logging.basicConfig(filename=path_log, level=logging.INFO, format='%(levelname)s : %(message)s')
    logging.info('############### Start Processing ###############')

    # To check result 
    result_list = []
    for result in os.listdir(dir_result):
        if ".npz" in result:
            result_name = f"{result.split('_')[0]}_{result.split('_')[1]}" 
            result_list.append(result_name)
    result_list = set(result_list)
    num_data_saved = 0
    number_list = os.listdir(dir_data)
    for number in number_list:
        # Pass if already saved
        if number in result_list:
            print('Data already saved: ' + number)
            num_data_saved += 1
            continue
            
        dir_dicom_final = None
        path_xml = None

        # find dicom directory
        # dir_number = os.path.join(dir_data, number)
        dir_number = dir_data + number + '/'
        directory_list = os.listdir(dir_number)

        if len(directory_list) == 0:
            print("Empty directory, number ", number)
            logging.info(f'Empty directory, number {number}')
            continue

        if directory_list[0][directory_list[0].rfind('.')+1:] == 'dcm':
            print("Only one dicom file, number ", number)
            logging.info(f'Only one dicom file, number {number}')
            continue
        for directory in directory_list:
            # print('directory: ', directory)
            # dir_dicoms = os.path.join(dir_number, directory)
            dir_dicoms = dir_number + directory + '/'
            dicom_list = os.listdir(dir_dicoms)
            for dicom in dicom_list:
                if dicom[dicom.rfind('.')+1:] == 'dcm':
                    print("Different directory structure, number ", number)
                    logging.info(f'Different directory structure, number {number}')
                    continue

                # dir_dicom = os.path.join(dir_dicoms, dicom)
                
                dir_dicom = dir_dicoms + dicom + '/'
                filename_list = os.listdir(dir_dicom)
                if len(filename_list) > 10:
                    for filename in filename_list:
                        if filename[:10] == 'defaultVOI':
                            _, fileNames, _ = get_itkImage_from_dicom(dir_dicom, flag_info=True)
                            if judge_axial(fileNames): 
                                dir_dicom_final = dir_dicom
                                dir_xml = os.path.join(dir_dicom, filename)
                                path_xml = os.path.join(dir_xml, 'VOIContour_0.xml')
                                break
                            elif judge_axial(fileNames) is None:
                                print("No iop tag, number ", number)
                                logging.info(f'No iop tag, number {number}')
                                break

                    if dir_dicom_final is not None: break
                else:
                    for filename in filename_list:
                        if '.' not in filename:
                            path = dir_dicom + filename + '/'
                            sub_filename_list = os.listdir(path)
                            if len(sub_filename_list) > 10:
                                for filename in sub_filename_list:
                                    if filename[:10] == 'defaultVOI':
                                        _, fileNames, _ = get_itkImage_from_dicom(path, flag_info=True)
                                        if judge_axial(fileNames): 
                                            dir_dicom_final = path
                                            dir_xml = os.path.join(path, filename)
                                            path_xml = os.path.join(dir_xml, 'VOIContour_0.xml')
                                            break
                                        elif judge_axial(fileNames) is None:
                                            print("No iop tag, number ", number)
                                            logging.info(f'No iop tag, number {number}')
                                            break

                                if dir_dicom_final is not None: break
                            else:
                                print("File name List = ", sub_filename_list )
                                for sub_filename in sub_filename_list:
                                    if '.' not in sub_filename:
                                        sub_path = path + sub_filename + '/'
                                        sub_filename_list_2 = os.listdir(sub_path)
                                        if len(sub_filename_list_2) > 10:
                                            for filename in sub_filename_list_2:
                                                if filename[:10] == 'defaultVOI':
                                                    _, fileNames, _ = get_itkImage_from_dicom(sub_path, flag_info=True)
                                                    if judge_axial(fileNames): 
                                                        dir_dicom_final = sub_path
                                                        dir_xml = os.path.join(sub_path, filename)
                                                        path_xml = os.path.join(dir_xml, 'VOIContour_0.xml')
                                                        break
                                                    elif judge_axial(fileNames) is None:
                                                        print("No iop tag, number ", number)
                                                        logging.info(f'No iop tag, number {number}')
                                                        break

                                            if dir_dicom_final is not None: break
            if dir_dicom_final is not None: break
        
        # print("dir_dicom_final: ", dir_dicom_final)
        if dir_dicom_final is None: 
            print("No dicom, number ", number)
            logging.info(f'No dicom, number {number}')
            continue
        else:   
            # Read dicom
            itkImage = get_itkImage_from_dicom(dir_dicom_final)

            # Get spacing and thickness from itkImage spacing([spacing, spacing, thickness])
            spacing_list_vector = itkImage.GetSpacing().GetVnlVector()
            spacing_list = itk.GetArrayFromVnlVector(spacing_list_vector)
            spacing_original = spacing_list[0]
            thickness_original = spacing_list[2]
            # print('spacing_original: ', spacing_original)
            # print('thickness_original: ', thickness_original)
            
            # # Get size
            region = itkImage.GetLargestPossibleRegion()
            size_image = region.GetSize()
            size_original = [size_image[2], size_image[0], size_image[1]]
            # # print(size_original)

            # Threshold ct: below = below, upper = 0 (itkImage)
            itkImage_threshold = ThresholdItkImage(itkImage, below=-150, upper=250)
           
            # Normalize (only ct)
            itkImage_normalized = NormalizeItkImage(itkImage_threshold, 1.0)

            arr_ct_procesed = itk.GetArrayFromImage(itkImage_normalized)

            # Read xml
            slice_number_list = list(range(size_original[0]))

            if os.path.isfile(path_xml) is False:
                print("No xml file, number ", number)
                logging.info(f'No xml file, number {number}')
                continue

            tree = parse(path_xml)
            root = tree.getroot()
            contours = root.findall("Contour")
            print(f"Process: {number}")
            for contour in contours:
                # get points
                slice_number = int(contour.findtext("Slice-number"))
                pt_list = [x.text for x in contour.iter('Pt')]
                if len(pt_list) != 4: 
                    print(f'pt_list is not 4 but {len(pt_list)}, slice_number={slice_number}, dir_dicom={dir_dicom_final}' )
                    logging.info(f'pt_list is not 4 but {len(pt_list)}, slice_number={slice_number}, dir_dicom={dir_dicom_final}' )
                    pt_list=None
                    break
                    # raise ValueError('pt_list is not 4 but %d' % len(pt_list))

                # get min and max points
                min_x, min_y = 9999, 9999
                max_x, max_y = -9999, -9999
                for pt in pt_list:
                    pt = pt.split(',')
                    x = float(pt[0])
                    y = float(pt[1])

                    if x <= min_x and y <= min_y:
                        min_x = x    
                        min_y = y   
                    elif x >= max_x and y >= max_y:
                        max_x = x    
                        max_y = y            

                if min_x == 9999 or min_y == 9999 or max_x == -9999 or max_y == -9999 : 
                    print(f'Wrong min or max point. min_x=={min_x}, min_y={min_y}, max_x={max_x}, max_y={max_y}. slice_number={slice_number}, dir_dicom={dir_dicom_final})')
                    logging.info(f'Wrong min or max point. min_x=={min_x}, min_y={min_y}, max_x={max_x}, max_y={max_y}. slice_number={slice_number}, dir_dicom={dir_dicom_final})')
                    min_x = None
                    break

                # heatmap
                if NUM_Landmark == 2:
                    ## landmark num = 2
                    landmarks = [(min_x, min_y),(max_x, max_y)]
                elif NUM_Landmark == 4:
                    ## landmark num = 4
                    landmarks = [(min_x, min_y), (min_x, max_y), (max_x, max_y), (max_x, min_y)]
                else:
                    print("Check landmark num")
                    exit()
                hms = _generate_hm_new(128, 128, landmarks, sigma=sigma)
                # save
                # title = number + '_' + dir_dicom_final[dir_dicom_final.rfind('')+1:] + '_' + str(slice_number)
                str_num = len(str(slice_number))
                zero_need = 4 - str_num
                z = ''
                for i in range(zero_need):
                    z += '0'
                new_name = f"{z}{slice_number}"
                title = f"{number}_{new_name}"
                # title = number + '_' + dir_dicom_final + '_' + str(slice_number)
                path_save = os.path.join(dir_result, title)  
                ct_save = arr_ct_procesed[int(slice_number)]

                np.savez_compressed(path_save, ct=ct_save, min_x=min_x, min_y=min_y, max_x=max_x, max_y=max_y, spacing_original=spacing_original, thickness_original=thickness_original, hms=hms)
                if slice_number in slice_number_list: slice_number_list.remove(slice_number)

                dir_image = os.path.join(dir_result, 'image')
                if os.path.isdir(dir_image) is False: os.mkdir(dir_image)
                # path_save = os.path.join(dir_image, title+'.jpg')
                # save_result_image(2, path_save, ct_save, hms, None)

            if pt_list is None: continue
            if min_x is None: continue

            # Save empty slices
            for slice_number in slice_number_list:
                str_num = len(str(slice_number))
                zero_need = 4 - str_num
                z = ''
                for i in range(zero_need):
                    z += '0'
                new_name = f"{z}{slice_number}"
                title = f"{number}_{new_name}_ZERO"
                path_save = os.path.join(dir_result, title)  
                ct_save = arr_ct_procesed[int(slice_number)]

                landmarks = [(0.0, 0.0),(0.0, 0.0)]
                hms = _generate_hm_new(128, 128, landmarks, sigma=sigma)

                np.savez_compressed(path_save, ct=ct_save, min_x=0.0, min_y=0.0, max_x=0.0, max_y=0.0, spacing_original=spacing_original, thickness_original=thickness_original, hms=hms)

            num_data_saved += 1
            print('Saved. %s' % number)
            print(f"{num_data_saved}/{len(number_list)}")
    
    print(f"Finished. {num_data_saved}/{len(number_list)}")
    logging.info(f"Finished. {num_data_saved}/{len(number_list)}")
    
def prepare_data_3slice(dir_data, dir_result, sigma):
    NUM_Landmark = 2
    if os.path.isdir(dir_result) is False: os.mkdir(dir_result)
    Check_COunt = 0
    # log
    path_log = os.path.join(dir_result, 'log.txt')
    logging.basicConfig(filename=path_log, level=logging.INFO, format='%(levelname)s : %(message)s')
    logging.info('############### Start Processing ###############')

    # To check result 
    result_list = os.listdir(dir_result)
    for i, result in enumerate(result_list):
        result_name = result[:result.find('_')]
        result_list[i] = result_name
    result_list = set(result_list)

    num_data_saved = 0
    number_list = os.listdir(dir_data)
    for number in number_list:
        # Pass if already saved
        if number in result_list:
            print('Data already saved: ' + number)
            num_data_saved += 1
            continue
            
        dir_dicom_final = None
        path_xml = None

        # find dicom directory
        # dir_number = os.path.join(dir_data, number)
        dir_number = dir_data + number + '/'
        directory_list = os.listdir(dir_number)

        if len(directory_list) == 0:
            print("Empty directory, number ", number)
            logging.info(f'Empty directory, number {number}')
            continue

        if directory_list[0][directory_list[0].rfind('.')+1:] == 'dcm':
            print("Only one dicom file, number ", number)
            logging.info(f'Only one dicom file, number {number}')
            continue
        for directory in directory_list:
            # print('directory: ', directory)
            # dir_dicoms = os.path.join(dir_number, directory)
            dir_dicoms = dir_number + directory + '/'
            dicom_list = os.listdir(dir_dicoms)
            for dicom in dicom_list:
                if dicom[dicom.rfind('.')+1:] == 'dcm':
                    print("Different directory structure, number ", number)
                    logging.info(f'Different directory structure, number {number}')
                    continue

                # dir_dicom = os.path.join(dir_dicoms, dicom)
                
                dir_dicom = dir_dicoms + dicom + '/'
                filename_list = os.listdir(dir_dicom)
                if len(filename_list) > 10:
                    for filename in filename_list:
                        if filename[:10] == 'defaultVOI':
                            _, fileNames, _ = get_itkImage_from_dicom(dir_dicom, flag_info=True)
                            if judge_axial(fileNames): 
                                dir_dicom_final = dir_dicom
                                dir_xml = os.path.join(dir_dicom, filename)
                                path_xml = os.path.join(dir_xml, 'VOIContour_0.xml')
                                break
                            elif judge_axial(fileNames) is None:
                                print("No iop tag, number ", number)
                                logging.info(f'No iop tag, number {number}')
                                break

                    if dir_dicom_final is not None: break
                else:
                    for filename in filename_list:
                        if '.' not in filename:
                            path = dir_dicom + filename + '/'
                            sub_filename_list = os.listdir(path)
                            if len(sub_filename_list) > 10:
                                for filename in sub_filename_list:
                                    if filename[:10] == 'defaultVOI':
                                        _, fileNames, _ = get_itkImage_from_dicom(path, flag_info=True)
                                        if judge_axial(fileNames): 
                                            dir_dicom_final = path
                                            dir_xml = os.path.join(path, filename)
                                            path_xml = os.path.join(dir_xml, 'VOIContour_0.xml')
                                            break
                                        elif judge_axial(fileNames) is None:
                                            print("No iop tag, number ", number)
                                            logging.info(f'No iop tag, number {number}')
                                            break

                                if dir_dicom_final is not None: break
                            else:
                                print("File name List = ", sub_filename_list )
                                for sub_filename in sub_filename_list:
                                    if '.' not in sub_filename:
                                        sub_path = path + sub_filename + '/'
                                        sub_filename_list_2 = os.listdir(sub_path)
                                        if len(sub_filename_list_2) > 10:
                                            for filename in sub_filename_list_2:
                                                if filename[:10] == 'defaultVOI':
                                                    _, fileNames, _ = get_itkImage_from_dicom(sub_path, flag_info=True)
                                                    if judge_axial(fileNames): 
                                                        dir_dicom_final = sub_path
                                                        dir_xml = os.path.join(sub_path, filename)
                                                        path_xml = os.path.join(dir_xml, 'VOIContour_0.xml')
                                                        break
                                                    elif judge_axial(fileNames) is None:
                                                        print("No iop tag, number ", number)
                                                        logging.info(f'No iop tag, number {number}')
                                                        break

                                            if dir_dicom_final is not None: break
            if dir_dicom_final is not None: break
        
        # print("dir_dicom_final: ", dir_dicom_final)
        if dir_dicom_final is None: 
            print("No dicom, number ", number)
            logging.info(f'No dicom, number {number}')
            continue
        else:   
            # Read dicom
            itkImage = get_itkImage_from_dicom(dir_dicom_final)

            # Get spacing and thickness from itkImage spacing([spacing, spacing, thickness])
            spacing_list_vector = itkImage.GetSpacing().GetVnlVector()
            spacing_list = itk.GetArrayFromVnlVector(spacing_list_vector)
            spacing_original = spacing_list[0]
            thickness_original = spacing_list[2]
            # print('spacing_original: ', spacing_original)
            # print('thickness_original: ', thickness_original)
            
            # # Get size
            region = itkImage.GetLargestPossibleRegion()
            size_image = region.GetSize()
            size_original = [size_image[2], size_image[0], size_image[1]]
            # # print(size_original)

            # Threshold ct: below = below, upper = 0 (itkImage)
            itkImage_threshold = ThresholdItkImage(itkImage, below=-150, upper=250)
           
            # Normalize (only ct)
            itkImage_normalized = NormalizeItkImage(itkImage_threshold, 1.0)

            arr_ct_procesed = itk.GetArrayFromImage(itkImage_normalized)

            # Read xml
            slice_number_list = list(range(size_original[0]))
            total_slice_num = len(slice_number_list)


            if os.path.isfile(path_xml) is False:
                print("No xml file, number ", number)
                logging.info(f'No xml file, number {number}')
                continue

            tree = parse(path_xml)
            root = tree.getroot()
            contours = root.findall("Contour")
            for contour in contours:
                # get points
                slice_number = int(contour.findtext("Slice-number"))
                pt_list = [x.text for x in contour.iter('Pt')]
                if len(pt_list) != 4: 
                    print(f'pt_list is not 4 but {len(pt_list)}, slice_number={slice_number}, dir_dicom={dir_dicom_final}' )
                    logging.info(f'pt_list is not 4 but {len(pt_list)}, slice_number={slice_number}, dir_dicom={dir_dicom_final}' )
                    pt_list=None
                    break
                    # raise ValueError('pt_list is not 4 but %d' % len(pt_list))

                # get min and max points
                min_x, min_y = 9999, 9999
                max_x, max_y = -9999, -9999
                for pt in pt_list:
                    pt = pt.split(',')
                    x = float(pt[0])
                    y = float(pt[1])

                    if x <= min_x and y <= min_y:
                        min_x = x    
                        min_y = y   
                    elif x >= max_x and y >= max_y:
                        max_x = x    
                        max_y = y            

                if min_x == 9999 or min_y == 9999 or max_x == -9999 or max_y == -9999 : 
                    print(f'Wrong min or max point. min_x=={min_x}, min_y={min_y}, max_x={max_x}, max_y={max_y}. slice_number={slice_number}, dir_dicom={dir_dicom_final})')
                    logging.info(f'Wrong min or max point. min_x=={min_x}, min_y={min_y}, max_x={max_x}, max_y={max_y}. slice_number={slice_number}, dir_dicom={dir_dicom_final})')
                    min_x = None
                    break

                # heatmap
                if NUM_Landmark == 2:
                    ## landmark num = 2
                    landmarks = [(min_x, min_y),(max_x, max_y)]
                elif NUM_Landmark == 4:
                    ## landmark num = 4
                    landmarks = [(min_x, min_y), (min_x, max_y), (max_x, max_y), (max_x, min_y)]
                else:
                    print("Check landmark num")
                    exit()
                hms = _generate_hm_new(128, 128, landmarks, sigma=sigma)
                # save
                # title = number + '_' + dir_dicom_final[dir_dicom_final.rfind('')+1:] + '_' + str(slice_number)
                str_num = len(str(slice_number))
                zero_need = 4 - str_num
                z = ''
                for i in range(zero_need):
                    z += '0'
                new_name = f"{z}{slice_number}"
                title = number + '_' + dir_dicom_final[dir_dicom_final.rfind('')+1:] + new_name
                # title = number + '_' + dir_dicom_final + '_' + str(slice_number)
                path_save = os.path.join(dir_result, title)  
                ct_save = []
                if slice_number in slice_number_list:
                    for idx in range(3):
                        if slice_number_list.index(slice_number) == 0:
                            s_num = int(slice_number)
                            if idx >= 1:
                                idx -= 1
                            ct_save.append(arr_ct_procesed[s_num + idx])
                        elif slice_number_list.index(slice_number) == (len(slice_number_list)-1):
                            s_num = int(slice_number) - 1
                            if idx == 2:
                                idx -= 1
                            ct_save.append(arr_ct_procesed[s_num + idx])
                        else:
                            s_num = int(slice_number) - 1
                            ct_save.append(arr_ct_procesed[s_num + idx])
                    slice_number_list.remove(slice_number)
                    np.savez_compressed(path_save, ct=ct_save, min_x=min_x, min_y=min_y, max_x=max_x, max_y=max_y, spacing_original=spacing_original, thickness_original=thickness_original, hms=hms)
                
            if pt_list is None: continue
            if min_x is None: continue

            # Save empty slices
            for slice_number in slice_number_list:
                str_num = len(str(slice_number))
                zero_need = 4 - str_num
                z = ''
                for i in range(zero_need):
                    z += '0'
                new_name = f"{z}{slice_number}"
                title = number + '_' + dir_dicom_final[dir_dicom_final.rfind('/')+1:] + new_name + '_ZERO'
                path_save = os.path.join(dir_result, title)  
                ct_save = arr_ct_procesed[int(slice_number)]

                landmarks = [(0.0, 0.0),(0.0, 0.0)]
                hms = _generate_hm_new(128, 128, landmarks, sigma=sigma)

                ct_save = []
                for idx in range(3):
                    if slice_number_list.index(slice_number) == 0:
                        s_num = int(slice_number)
                        if idx >= 1:
                            idx -= 1
                        ct_save.append(arr_ct_procesed[s_num + idx])
                    elif slice_number_list.index(slice_number) == (len(slice_number_list)-1):
                        s_num = int(slice_number) - 1
                        if idx == 2:
                            idx -= 1
                        ct_save.append(arr_ct_procesed[s_num + idx])
                    else:
                        s_num = int(slice_number) - 1
                        ct_save.append(arr_ct_procesed[s_num + idx])

                np.savez_compressed(path_save, ct=ct_save, min_x=0.0, min_y=0.0, max_x=0.0, max_y=0.0, spacing_original=spacing_original, thickness_original=thickness_original, hms=hms)

            num_data_saved += 1
            print('Saved. %s' % number)
            print(f"{num_data_saved}/{len(number_list)}")
    
    print(f"Finished. {num_data_saved}/{len(number_list)}")
    logging.info(f"Finished. {num_data_saved}/{len(number_list)}")
  
def prepare_data_val(dir_data, dir_result, sigma, num_landmark):
    NUM_Landmark = num_landmark
    if os.path.isdir(dir_result) is False: os.mkdir(dir_result)

    # log
    path_log = os.path.join(dir_result, 'log.txt')
    logging.basicConfig(filename=path_log, level=logging.INFO, format='%(levelname)s : %(message)s')
    logging.info('############### Start Processing ###############')

    # To check result 
    result_list = []
    for result in os.listdir(dir_result):
        if "ANO_" in result:
            result_list.append(f"{result.split('_')[0]}_{result.split('_')[1]}")
    if len(result_list) > 1:
        result_list = set(result_list)
    print("already data num : ", len(result_list))
    num_data_saved = 0
    # number_list = os.listdir(dir_data)
    number_list = [x for x in os.listdir(dir_data) if 'ANO_' in x]
    for number in number_list:
        # print("Number: ", number)
        
        # Pass if already saved
        if number in result_list:
            print('Data already saved: ' + number)
            num_data_saved += 1
            continue
            
        dir_dicom_final = None
        path_xml = None

        # find dicom directory
        # dir_number = os.path.join(dir_data, number)
        dir_number = dir_data + number + '/'
        directory_list = os.listdir(dir_number)

        if len(directory_list) == 0:
            print("Empty directory, number ", number)
            logging.info(f'Empty directory, number {number}')
            continue

        if directory_list[0][directory_list[0].rfind('.')+1:] == 'dcm':
            print("Only one dicom file, number ", number)
            logging.info(f'Only one dicom file, number {number}')
            continue
        for directory in directory_list:
            # dir_dicoms = os.path.join(dir_number, directory)
            dir_dicom = ''
            filename_list = ''
            if "axial" in directory_list:
                dir_dicom = dir_number + '/axial/'
                filename_list = os.listdir(dir_dicom)
            else:
                dir_dicom = dir_number + directory + '/axial/'
                filename_list = os.listdir(dir_dicom)
            if len(filename_list) > 10:
                dir_dicom_final = dir_dicom
                for filename in filename_list:
                    if "VOIContour_0" in filename:
                        # _, fileNames, _ = get_itkImage_from_dicom(dir_dicom, flag_info=True)
                        path_xml = os.path.join(dir_dicom, 'VOIContour_0.xml')
                    
                    if "Q" in filename:
                        dir_dicom = f"{dir_dicom}Q/"
                        for voi in os.listdir(dir_dicom):
                            if "VOIContour_0" in voi:
                                path_xml = os.path.join(dir_dicom, 'VOIContour_0.xml')
                    
                if dir_dicom_final is not None: break
            else:
                print(dir_dicom)
                exit()
            if dir_dicom_final is not None: break
        
        # print("dir_dicom_final: ", dir_dicom_final)
        if path_xml is None: 

            print("No VOI, number ", number)
            logging.info(f'No VOI, number {number}')
            
            # Read dicom
            itkImage = get_itkImage_from_dicom(dir_dicom_final)

            # Get spacing and thickness from itkImage spacing([spacing, spacing, thickness])
            spacing_list_vector = itkImage.GetSpacing().GetVnlVector()
            spacing_list = itk.GetArrayFromVnlVector(spacing_list_vector)
            spacing_original = spacing_list[0]
            thickness_original = spacing_list[2]
            # print('spacing_original: ', spacing_original)
            # print('thickness_original: ', thickness_original)
            
            # # Get size
            region = itkImage.GetLargestPossibleRegion()
            size_image = region.GetSize()
            size_original = [size_image[2], size_image[0], size_image[1]]
            # # print(size_original)

            # Threshold ct: below = below, upper = 0 (itkImage)
            itkImage_threshold = ThresholdItkImage(itkImage, below=-150, upper=250)
           
            # Normalize (only ct)
            itkImage_normalized = NormalizeItkImage(itkImage_threshold, 1.0)

            arr_ct_procesed = itk.GetArrayFromImage(itkImage_normalized)

            # Read xml
            slice_number_list = list(range(size_original[0]))

            # Save empty slices
            for slice_number in slice_number_list:
                title = number + '_' + dir_dicom_final[dir_dicom_final.rfind('/')+1:] + str(slice_number) + '_ZERO'
                path_save = os.path.join(dir_result, title)  
                ct_save = arr_ct_procesed[int(slice_number)]

                landmarks = [(0.0, 0.0),(0.0, 0.0)]
                hms = _generate_hm_new(128, 128, landmarks, sigma=sigma)

                np.savez_compressed(path_save, ct=ct_save, min_x=0.0, min_y=0.0, max_x=0.0, max_y=0.0, spacing_original=spacing_original, thickness_original=thickness_original, hms=hms)

            num_data_saved += 1
            print('Saved. %s' % number)
            print(f"{num_data_saved}/{len(number_list)}")
            
        else:   
            # Read dicom
            itkImage = get_itkImage_from_dicom(dir_dicom_final)

            # Get spacing and thickness from itkImage spacing([spacing, spacing, thickness])
            spacing_list_vector = itkImage.GetSpacing().GetVnlVector()
            spacing_list = itk.GetArrayFromVnlVector(spacing_list_vector)
            spacing_original = spacing_list[0]
            thickness_original = spacing_list[2]
            # print('spacing_original: ', spacing_original)
            # print('thickness_original: ', thickness_original)
            
            # # Get size
            region = itkImage.GetLargestPossibleRegion()
            size_image = region.GetSize()
            size_original = [size_image[2], size_image[0], size_image[1]]
            # # print(size_original)

            # Threshold ct: below = below, upper = 0 (itkImage)
            itkImage_threshold = ThresholdItkImage(itkImage, below=-150, upper=250)
           
            # Normalize (only ct)
            itkImage_normalized = NormalizeItkImage(itkImage_threshold, 1.0)

            arr_ct_procesed = itk.GetArrayFromImage(itkImage_normalized)

            # Read xml
            slice_number_list = list(range(size_original[0]))

            if os.path.isfile(path_xml) is False:
                print("No xml file, number ", number)
                logging.info(f'No xml file, number {number}')
                continue

            tree = parse(path_xml)
            root = tree.getroot()
            contours = root.findall("Contour")
            for contour in contours:
                # get points
                slice_number = int(contour.findtext("Slice-number"))
                pt_list = [x.text for x in contour.iter('Pt')]
                if len(pt_list) != 4: 
                    print(f'pt_list is not 4 but {len(pt_list)}, slice_number={slice_number}, dir_dicom={dir_dicom_final}' )
                    logging.info(f'pt_list is not 4 but {len(pt_list)}, slice_number={slice_number}, dir_dicom={dir_dicom_final}' )
                    pt_list=None
                    break
                    # raise ValueError('pt_list is not 4 but %d' % len(pt_list))

                # get min and max points
                min_x, min_y = 9999, 9999
                max_x, max_y = -9999, -9999
                for pt in pt_list:
                    pt = pt.split(',')
                    x = float(pt[0])
                    y = float(pt[1])

                    if x <= min_x and y <= min_y:
                        min_x = x    
                        min_y = y   
                    elif x >= max_x and y >= max_y:
                        max_x = x    
                        max_y = y            

                if min_x == 9999 or min_y == 9999 or max_x == -9999 or max_y == -9999 : 
                    print(f'Wrong min or max point. min_x=={min_x}, min_y={min_y}, max_x={max_x}, max_y={max_y}. slice_number={slice_number}, dir_dicom={dir_dicom_final})')
                    logging.info(f'Wrong min or max point. min_x=={min_x}, min_y={min_y}, max_x={max_x}, max_y={max_y}. slice_number={slice_number}, dir_dicom={dir_dicom_final})')
                    min_x = None
                    break

                # heatmap
                if NUM_Landmark == 2:
                    ## landmark num = 2
                    landmarks = [(min_x, min_y),(max_x, max_y)]
                elif NUM_Landmark == 4:
                    ## landmark num = 4
                    landmarks = [(min_x, min_y), (min_x, max_y), (max_x, max_y), (max_x, min_y)]
                else:
                    print("Check landmark num")
                    exit()
                hms = _generate_hm_new(128, 128, landmarks, sigma=sigma)
                # save
                # title = number + '_' + dir_dicom_final[dir_dicom_final.rfind('')+1:] + '_' + str(slice_number)
                title = number + '_' + dir_dicom_final[dir_dicom_final.rfind('')+1:] + str(slice_number)
                # title = number + '_' + dir_dicom_final + '_' + str(slice_number)
                path_save = os.path.join(dir_result, title)  
                ct_save = arr_ct_procesed[int(slice_number)]

                np.savez_compressed(path_save, ct=ct_save, min_x=min_x, min_y=min_y, max_x=max_x, max_y=max_y, spacing_original=spacing_original, thickness_original=thickness_original, hms=hms)
                if slice_number in slice_number_list: slice_number_list.remove(slice_number)

            if pt_list is None: continue
            if min_x is None: continue

            # Save empty slices
            for slice_number in slice_number_list:
                title = number + '_' + dir_dicom_final[dir_dicom_final.rfind('/')+1:] + str(slice_number) + '_ZERO'
                path_save = os.path.join(dir_result, title)  
                ct_save = arr_ct_procesed[int(slice_number)]

                landmarks = [(0.0, 0.0),(0.0, 0.0)]
                hms = _generate_hm_new(128, 128, landmarks, sigma=sigma)

                np.savez_compressed(path_save, ct=ct_save, min_x=0.0, min_y=0.0, max_x=0.0, max_y=0.0, spacing_original=spacing_original, thickness_original=thickness_original, hms=hms)

            num_data_saved += 1
            print('Saved. %s' % number)
            print(f"{num_data_saved}/{len(number_list)}")
    
    print(f"Finished. {num_data_saved}/{len(number_list)}")
    logging.info(f"Finished. {num_data_saved}/{len(number_list)}")

def Render3DNumpyArray(arr, spacing=1, thickness=1, path_save=None):
    
    VTK_data = numpy_to_vtk(num_array=arr.ravel(), deep=True, array_type=vtk.VTK_FLOAT)    
    vtkImage = vtk.vtkImageData()
    vtkImage.SetOrigin([0, 0, 0])
    vtkImage.SetDimensions(arr.shape[2], arr.shape[1], arr.shape[0])
    vtkImage.AllocateScalars(vtk.VTK_FLOAT, 1)
    vtkImage.SetSpacing([spacing,spacing,thickness])
    vtkImage.GetPointData().SetScalars(VTK_data)

    alphaChannelFunc = vtk.vtkPiecewiseFunction()
    colorFunc = vtk.vtkColorTransferFunction()

    # alphaChannelFunc.AddPoint(np.min(arr), 0.8)
    alphaChannelFunc.AddPoint(0, 0.0)
    alphaChannelFunc.AddPoint(np.max(arr), 0.5)

    # colorFunc.AddRGBPoint(np.min(arr), 70/255, 150/255, 220/255)
    # colorFunc.AddRGBPoint(np.max(arr), 70/255, 150/255, 220/255)

    colorFunc.AddRGBPoint(np.min(arr), 0, 1, 0)
    colorFunc.AddRGBPoint(np.max(arr), 1, 0, 0)

    volumeProperty = vtk.vtkVolumeProperty()
    volumeProperty.SetColor(colorFunc)
    volumeProperty.SetScalarOpacity(alphaChannelFunc)
    volumeProperty.ShadeOn()

    mapper = vtk.vtkSmartVolumeMapper()
    mapper.SetInputData(vtkImage)

    actor = vtk.vtkVolume()
    actor.SetProperty(volumeProperty)
    actor.SetMapper(mapper)
    actor.RotateX(-70.0)
    # actor.RotateY(15.0)
    # bounds = actor.GetBounds()
    # mid_x = (bounds[1]-bounds[0])/2
    # mid_y = (bounds[3]-bounds[2])/2
    # mid_z = (bounds[5]-bounds[4])/2

    # source = vtk.vtkSphereSource()
    # source.SetCenter(mid_x, mid_y, mid_z)
    # source.SetRadius(5.0)
    # mapper_source = vtk.vtkPolyDataMapper()
    # mapper_source.SetInputConnection(source.GetOutputPort())
    # actor_source = vtk.vtkActor()
    # actor_source.SetMapper(mapper_source)

    # source_0 = vtk.vtkSphereSource()
    # source_0.SetCenter(0, 0, 0)
    # source_0.SetRadius(5.0)
    # mapper_source_0 = vtk.vtkPolyDataMapper()
    # mapper_source_0.SetInputConnection(source_0.GetOutputPort())
    # actor_source_0 = vtk.vtkActor()
    # actor_source_0.SetMapper(mapper_source_0)

    # camera = vtk.vtkCamera()
    # camera.SetPosition(mid_x,mid_y,mid_z+100)
    # camera.SetFocalPoint(mid_x, mid_y, mid_z)
   
    renderer = vtk.vtkRenderer()
    renderer.SetBackground(1,1,1)
    renderer.AddVolume(actor)
    # renderer.SetActiveCamera(camera)

    # renderer.AddActor(actor_source)
    # renderer.AddActor(actor_source_0)

    renderWindow = vtk.vtkRenderWindow()
    renderWindow.AddRenderer(renderer)
    renderWindowInteractor = vtk.vtkRenderWindowInteractor()
    renderWindowInteractor.SetInteractorStyle(vtk.vtkInteractorStyleTrackballCamera())
    renderWindowInteractor.SetRenderWindow(renderWindow)
    if path_save is None : renderWindowInteractor.Start()
    else :  
        renderWindow.Render()

        windowToImageFilter = vtk.vtkWindowToImageFilter()
        windowToImageFilter.SetInput(renderWindow)
        # windowToImageFilter.SetMagnification(3) #set the resolution of the output image (3 times the current resolution of vtk render window)
        windowToImageFilter.SetInputBufferTypeToRGBA() #also record the alpha (transparency) channel
        windowToImageFilter.ReadFrontBufferOff() #read from the back buffer
        windowToImageFilter.Update()

        writer = vtk.vtkPNGWriter()
        writer.SetFileName(path_save)
        writer.SetInputConnection(windowToImageFilter.GetOutputPort())
        writer.Write()

def GetDataInformationFromDicom(dir_data):
    # Init excel
    dir_parent = os.path.abspath(os.path.join(dir_data, os.pardir))
    path_excel = os.path.join(dir_parent, 'information.xlsx')
    wb = openpyxl.Workbook()
    worksheet = wb.active
    worksheet.append(["number", "dicom", "spacing", "thickness", "num_slice", "size", "min", "max", "modality", "manufacturer", "rescale_intercept", "rescale_slope", "window_center", "window_level"])   
    wb.save(path_excel)

    number_list = os.listdir(dir_data)
    for number in number_list:
        # print("Number: ", number)
        
        dir_dicom_final = None
        dicom_final = None
        path_xml = None

        # find dicom directory
        dir_number = os.path.join(dir_data, number)
        directory_list = os.listdir(dir_number)

        if len(directory_list) == 0:
            wb = openpyxl.load_workbook(path_excel)
            ws = wb.active
            ws.append([number, 'Empty directory'])
            wb.save(path_excel)
            
            print("Empty directory, number ", number)
            continue

        if directory_list[0][directory_list[0].rfind('.')+1:] == 'dcm':
            wb = openpyxl.load_workbook(path_excel)
            ws = wb.active
            ws.append([number, 'Only one dicom file'])
            wb.save(path_excel)

            print("Only one dicom file, number ", number)
            continue

        for directory in directory_list:
            # print('directory: ', directory)
            dir_dicoms = os.path.join(dir_number, directory)
            dicom_list = os.listdir(dir_dicoms)
            for dicom in dicom_list:

                if dicom[dicom.rfind('.')+1:] == 'dcm':
                    wb = openpyxl.load_workbook(path_excel)
                    ws = wb.active
                    ws.append([number, 'Different directory structure'])
                    wb.save(path_excel)

                    print("Different directory structure, number ", number)
                    continue

                dir_dicom = os.path.join(dir_dicoms, dicom)
                filename_list = os.listdir(dir_dicom)
                if len(filename_list) > 10:
                    for filename in filename_list:
                        if filename[:10] == 'defaultVOI':
                            _, fileNames, _ = get_itkImage_from_dicom(dir_dicom, flag_info=True)
                            if judge_axial(fileNames): 
                                dir_dicom_final = dir_dicom
                                dicom_final = dicom
                                dir_xml = os.path.join(dir_dicom, filename)
                                path_xml = os.path.join(dir_xml, 'VOIContour_0.xml')
                                break
                            elif judge_axial is None:
                                wb = openpyxl.load_workbook(path_excel)
                                ws = wb.active
                                ws.append([number, 'No iop tag'])
                                wb.save(path_excel)

                                print("No iop tag, number ", number)
                                break
                    if dir_dicom_final is not None: break
                else:
                    for filename in filename_list:
                        if '.' not in filename:
                            path = dir_dicom + filename + '/'
                            sub_filename_list = os.listdir(path)
                            if len(sub_filename_list) > 10:
                                for filename in sub_filename_list:
                                    if filename[:10] == 'defaultVOI':
                                        _, fileNames, _ = get_itkImage_from_dicom(path, flag_info=True)
                                        if judge_axial(fileNames): 
                                            dir_dicom_final = path
                                            dir_xml = os.path.join(path, filename)
                                            path_xml = os.path.join(dir_xml, 'VOIContour_0.xml')
                                            break
                                        elif judge_axial(fileNames) is None:
                                            print("No iop tag, number ", number)
                                            logging.info(f'No iop tag, number {number}')
                                            break

                                if dir_dicom_final is not None: break
                            else:
                                print("File name List = ", sub_filename_list )
                                for sub_filename in sub_filename_list:
                                    if '.' not in sub_filename:
                                        sub_path = path + sub_filename + '/'
                                        sub_filename_list_2 = os.listdir(sub_path)
                                        if len(sub_filename_list_2) > 10:
                                            for filename in sub_filename_list_2:
                                                if filename[:10] == 'defaultVOI':
                                                    _, fileNames, _ = get_itkImage_from_dicom(sub_path, flag_info=True)
                                                    if judge_axial(fileNames): 
                                                        dir_dicom_final = sub_path
                                                        dir_xml = os.path.join(sub_path, filename)
                                                        path_xml = os.path.join(dir_xml, 'VOIContour_0.xml')
                                                        break
                                                    elif judge_axial(fileNames) is None:
                                                        print("No iop tag, number ", number)
                                                        logging.info(f'No iop tag, number {number}')
                                                        break

                                            if dir_dicom_final is not None: break
            if dir_dicom_final is not None: break
        
        # print("dir_dicom_final: ", dir_dicom_final)
     
        if dir_dicom_final is None: 
            wb = openpyxl.load_workbook(path_excel)
            ws = wb.active
            ws.append([number, 'No dicom'])
            wb.save(path_excel)

            print("No dicom, number ", number)
            continue
        else:   
            # Read dicom
            itkImage_ct, reader = get_itkImage_from_dicom(dir_dicom_final, flag_meta=True)
            metadata = reader.GetImageIO().GetMetaDataDictionary()

            # Get spacing and thickness from itkImage spacing([spacing, spacing, thickness])
            spacing_list_vector = itkImage_ct.GetSpacing().GetVnlVector()
            spacing_list = itk.GetArrayFromVnlVector(spacing_list_vector)
            spacing_original = spacing_list[0]
            thickness_original = spacing_list[2]

            # Get size
            arr_original = itk.GetArrayFromImage(itkImage_ct)
            size_original = np.copy(arr_original.shape)

            # Get modality
            try: modality = metadata["0008|0060"]
            except: modality = 'NULL'
            
            # Get manufacturer 
            try: manufacturer = metadata["0008|1090"]
            except: manufacturer = 'NULL'

            # Get RescaleIntercept, RescaleSlope
            try:
                rescale_intercept = metadata["0028|1052"]
                rescale_slope = metadata["0028|1053"]
            except:
                rescale_intercept = 'NULL'
                rescale_slope = 'NULL'
            
            # Get WindowCenter, WindowLevel
            try:
                window_center = metadata["0028|1050"] 
                window_level = metadata["0028|1051"]
            except:
                window_center = 'NULL'
                window_level = 'NULL'

            # Save at excel
            wb = openpyxl.load_workbook(path_excel)
            ws = wb.active
            #worksheet.append(["number", "dicom", "spacing", "thickness", "num_slice", "size", "min", "max", "modality", "manufacturer", "rescale_intercept", "rescale_slope", "window_center", "window_level"])   
            ws.append([number, dir_dicom_final, spacing_original, thickness_original, size_original[0], str(size_original[1])+','+str(size_original[2]), np.min(arr_original), np.max(arr_original), modality, manufacturer, rescale_intercept, rescale_slope, window_center, window_level])
            wb.save(path_excel)
            
            print("Saved, number ", number)

def add_heatmap(dir_data, dir_result, sigma):
    '''
    이미 저장되어 있는 npz 불러와서 heatmap 추가 저장
    '''
    if os.path.isdir(dir_result) is False: os.mkdir(dir_result)
    
    dir_image = os.path.join(dir_result, 'image')
    if os.path.isdir(dir_image) is False: os.mkdir(dir_image)
    
    data_list = os.listdir(dir_data)
    for name in data_list:      
        path_data = os.path.join(dir_data, name)
        ct, min_x, min_y, max_x, max_y, spacing_original, thickness_original = load_npz(path_data)

        landmarks = [(min_x, min_y),(max_x, max_y)]
        hms, aas = _generate_hm_new(128, 128, landmarks, sigma=sigma)

        path_save = os.path.join(dir_result, name)
        np.savez_compressed(path_save, ct=ct, min_x=min_x, min_y=min_y, max_x=max_x, max_y=max_y, spacing_original=spacing_original, thickness_original=thickness_original, hms=hms)

        # Debug image save
        name = name[:name.rfind('.')]
        for i, (hm, a) in enumerate(zip(hms, aas)):
            path_zoom = os.path.join(dir_image, name+'_'+str(i)+'_zoom.png')
            fig = plt.figure(0)
            plt.imshow(hm[a[0]:a[1], a[2]:a[3]])
            plt.savefig(path_zoom)
            plt.close()

            path_all = os.path.join(dir_image, name+'_'+str(i)+'_all.png')
            fig = plt.figure(0)
            plt.imshow(hm)
            plt.savefig(path_all)
            plt.close()

        print(path_save)

def _generate_hm_new(height, width, landmarks, sigma=3, divide_factor = 4): # sigma -> 랜드마크 기준 얼마나 퍼진 heatmap 인지
    hms = np.zeros(shape=(len(landmarks), height, width), dtype=np.float32)
    # print(hms.shape)

    for i in range(len(landmarks)):
        if (landmarks[i][0] != -1.0) and (landmarks[i][1] != -1.0):
            x, y = round(landmarks[i][0]/divide_factor), round(landmarks[i][1]/divide_factor)
            if x < 0 or y < 0 or x >= width or y >= height:
                continue
            ul = int(x - 3 * sigma - 1), int(y - 3 * sigma - 1)
            br = int(x + 3 * sigma + 2), int(y + 3 * sigma + 2)

            c, d = max(0, -ul[0]), min(br[0], width) - ul[0]
            a, b = max(0, -ul[1]), min(br[1], height) - ul[1]

            cc, dd = max(0, ul[0]), min(br[0], width)
            aa, bb = max(0, ul[1]), min(br[1], height)
            gaussian = np.maximum(hms[i, aa:bb, cc:dd], _makeGaussian_new(sigma)[a:b, c:d])
            hms[i, aa:bb, cc:dd] = gaussian

            # print(a, b, c, d)
            # print(aa, bb, cc, dd)

            # fig = plt.figure(0)
            # plt.imshow(hms[i, aa:bb, cc:dd])
            # plt.show()

            # fig = plt.figure(0)
            # plt.imshow(hms[i])
            # plt.show()

    return hms

def _makeGaussian_new(sigma=3):
    size = 6 * sigma + 3

    x = np.arange(0, size, 1, float)
    y = x[:, np.newaxis]

    x0, y0 = 3 * sigma + 1, 3 * sigma + 1
    gaussian = np.exp(- ((x - x0) ** 2 + (y - y0) ** 2) / (2 * sigma ** 2))
    return gaussian

def load_npz(path_file):
    '''
    Load data from npz file
    '''
    ct, min_x, min_y, max_x, max_y, hms, spacing_original, thickness_original = None, None, None, None, None, None, None, None
    data = np.load(path_file, allow_pickle=True)
    if 'ct' in data.files: ct = data['ct'].astype(np.float32)
    if 'min_x' in data.files: min_x = float(data['min_x'])
    if 'min_y' in data.files: min_y = float(data['min_y'])
    if 'max_x' in data.files: max_x = float(data['max_x'])
    if 'max_y' in data.files: max_y = float(data['max_y'])
    if 'hms' in data.files: hms = data['hms']
    if 'spacing_original' in data.files: spacing_original = float(data['spacing_original'])
    if 'thickness_original' in data.files: thickness_original = float(data['thickness_original'])

    del data.f
    data.close()

    # print("Data loaded, ", path_file)
    return ct, min_x, min_y, max_x, max_y, hms, spacing_original, thickness_original

def return_class_list(dir_data):
    list_dir = os.listdir(dir_data)
    patient_list = []
    for num in range(len(list_dir)):
        patient = list_dir[num].split('_')[1]
        patient_list.append(patient)

    patient_list = sorted(list(set(patient_list)))
    
    list_zero, list_nonzero = [], []
    for name in list_dir:
        if "ZERO" in name:
            list_zero.append(name)
        else: 
            list_nonzero.append(name)
    result_list = []
    if len(list_zero) == len(list_nonzero):
        result_list = [*sum(zip(list_zero, list_nonzero),())]
    else:
        result_list = sorted(list_zero + list_nonzero)
    return result_list
            
def divide_data(dir_data, ratio_val, ratio_test):
    """
    Divide data into train,val,test folder
    dir_data: directory of folder containing npz files
    """

    # Calculate numbers
    list_class_number = [] #[(num_train, num_val, num_test), (...), ...]
    lists_class = return_class_list(dir_data)
    for list_class in lists_class:
        num_class = len(list_class)
        num_val = int(num_class * ratio_val)
        num_test = int(num_class * ratio_test)
        num_train = num_class - (num_val + num_test)
        list_class_number.append((num_train, num_val, num_test))

def divide_data_patient(dir_data, ratio_test, ratio_val):
    """
    Divide data into train,val,test folder
    dir_data: directory of folder containing npz files
    """
    # To check result 
    result_list = os.listdir(dir_data)
    patient_list = []
    all_file_list = []
    for i, result in enumerate(result_list):
        if '.npz' in result:
            all_file_list.append(result)
            result_name = result.split("_")[1]
            patient_list.append(result_name)
    patient_list = set(patient_list)
    patient_num = len(patient_list)
    print("Total File Num = ", patient_num)
    test_num = int(patient_num * ratio_test)
    val_num = int(patient_num * ratio_val)
    trian_num = patient_num - test_num - val_num
    test_list = random.sample(patient_list, test_num)
    for test in test_list:
        patient_list.remove(test)
    val_list = random.sample(patient_list, val_num)
    for val in val_list:
        patient_list.remove(val)
    train_list = patient_list
    print("Train File Num = ", len(train_list))
    print("Test File Num = ", len(test_list))
    print("Val File Num = ", len(val_list))
    print("")
    dir_train = dir_data + "train/"
    os.makedirs(dir_train, exist_ok=True)
    dir_test = dir_data + "test/"
    os.makedirs(dir_test, exist_ok=True)
    dir_val = dir_data + "val/"
    os.makedirs(dir_val, exist_ok=True) 

    for file_name in all_file_list:
        print(f"Progress -------------  {all_file_list.index(file_name)+1} / {len(all_file_list)}", end='\r')
        patient_num = file_name.split("_")[1]
        if patient_num in train_list:
            ori_path = dir_data + file_name
            new_path = dir_train + file_name
            shutil.move(ori_path, new_path)
        elif patient_num in test_list:
            ori_path = dir_data + file_name
            new_path = dir_test + file_name
            shutil.move(ori_path, new_path)
        elif patient_num in val_list:
            ori_path = dir_data + file_name
            new_path = dir_val + file_name
            shutil.move(ori_path, new_path)
        else:
            print("Check Data")
            print(file_name)
            exit()
    print("")
    print("Finish")

def return_divide(dir_data):
    '''
    train,test,val로 나누어진 파일들 다시 합치기
    '''
    dir_train = os.path.join(dir_data, 'train')
    dir_val = os.path.join(dir_data, 'val')
    dir_test = os.path.join(dir_data, 'test')

    directories = [dir_train, dir_val, dir_test]
    for directory in directories:
        dir_list = os.listdir(directory)
        for name in dir_list:
            path = os.path.join(directory, name)
            path_new = os.path.join(dir_data, name)
            shutil.move(path, path_new)
        shutil.rmtree(directory)
       
def convert_heatmap_to_landmark(heatmap):
    return np.flip(np.array(np.unravel_index(heatmap.argmax(), heatmap.shape)))

def judge_zero(number):
    # zero = number[number.rfind('__')+2:]
    # if zero == 'ZERO': return True
    # else: return False
    if "ZERO" in number:
        return True
    else:
        return False

def divide_year(dir_data):
    dir_2013 = os.path.join(dir_data, '2013')
    if os.path.isdir(dir_2013) is False: os.mkdir(dir_2013)
    dir_2014 = os.path.join(dir_data, '2014')
    if os.path.isdir(dir_2014) is False: os.mkdir(dir_2014)

    list_dir = os.listdir(dir_data)
    for name in list_dir:
        path_old = os.path.join(dir_data, name)
        year = name[:name.find('_')]
        if year == '2013':
            path_new = os.path.join(dir_2013, name)
            shutil.move(path_old, path_new)
        elif year == '2014':
            path_new = os.path.join(dir_2014, name)
            shutil.move(path_old, path_new)
        print(name)

def divide_balance(dir_data):
    zero_data = dir_data[:-1] + "_Zero/"
    os.makedirs(zero_data, exist_ok=True)

    train_list = os.listdir(dir_data)
    patients_list = []
    for num in range(len(train_list)):
        patients_list.append(train_list[num].split('_')[1])
    patients_list = list(set(patients_list))
    patients_list.sort()

    Total_Tumor_List = []
    Total_Zero_List = []
    for patient in patients_list:
        tumor_list = []
        zero_list = []
        for num in range(len(train_list)):
            if patient in train_list[num]:
                if "ZERO" not in train_list[num]:
                    tumor_list.append(train_list[num])
                else:
                    zero_list.append(train_list[num])
        Total_Tumor_List.append(tumor_list)
        Total_Zero_List.append(zero_list)        

    Result_List = []
    for t in range(len(Total_Tumor_List)):
        tumor_list = sorted(Total_Tumor_List[t]) 
        patient = tumor_list[0].split('_')[1]
        start = int(tumor_list[0].split('_')[2][:-4])
        tumor_num = len(tumor_list)
        result_zero = []
        for z in range(len(Total_Zero_List)):
            if patient in Total_Zero_List[z][0]:
                zero_list = sorted(Total_Zero_List[z])
                rand_zero_list = random.sample(zero_list, tumor_num)
                
                for num in range(len(zero_list)):
                    file_name = zero_list[num]
                    if file_name not in rand_zero_list:
                        shutil.move(dir_data + file_name, zero_data + file_name)

def copy_file(ori_path, copy_path):
    print(f"{ori_path} Copy TUMOR FILE")
    tumor_list = []
    zero_list = []
    if "LN2_2014_HU_3slice" in ori_path:
        print(f"{ori_path} Copy ZERO FILE")
    for name in os.listdir(ori_path):
        if "ZERO" not in name:
            tumor_list.append(name)
        if "LN2_2014_HU_3slice" in ori_path:
            if "ZERO" in name:
                zero_list.append(name)
    for num in range(len(tumor_list)):
        shutil.copy(ori_path + tumor_list[num], copy_path + tumor_list[num])
        print(f"Tumor copy Progres -------- {num}/{len(tumor_list)}", end='\r')
    for num in range(len(zero_list)):
        shutil.copy(ori_path + zero_list[num], copy_path + zero_list[num])
        print(f"Zero copy Progres -------- {num}/{len(zero_list)}", end='\r')

def move_file(ori_path, move_path):
    file_list = []
    for name in os.listdir(ori_path):
        file_list.append(name)

    for num in range(len(file_list)):
        shutil.move(ori_path + file_list[num], move_path + file_list[num])
        print(f"Progres -------- {num}/{len(file_list)}", end='\r')

def count_patient_num(dir_path):
    file_list = []
    for file_name in os.listdir(dir_path):
        patient = file_name.split('_')[1]
        file_list.append(patient)

    file_list = list(set(file_list))
    print(len(file_list))

def rename_slice(dir_path):
    new_name_list = []
    ori_name_list = []
    for file_name in os.listdir(dir_path):
        ori_name_list.append(file_name)
        file_split = file_name.split('_')
        if len(file_split) == 3:     
            str_num = len(file_split[2][:-4])
            zero_need = 4 - str_num
            z = ''
            for i in range(zero_need):
                z += '0'
            new_num = z + file_split[2][:-4]   
            new_name = f"{file_split[0]}_{file_split[1]}_{new_num}.npz"   
            new_name_list.append(new_name)
        else:
            str_num = len(file_split[2])
            zero_need = 4 - str_num
            z = ''
            for i in range(zero_need):
                z += '0'
            new_num = z + file_split[2]
            new_num = z + file_split[2]
            new_name = f"{file_split[0]}_{file_split[1]}_{new_num}_{file_split[3]}"
            new_name_list.append(new_name)
    
    for num in range(len(ori_name_list)):
        os.rename(f"{dir_path}{ori_name_list[num]}", f"{dir_path}{new_name_list[num]}")
        print(f"Progres -------- {num+1}/{len(ori_name_list)}", end='\r')

def save_result_image(path_save, input_tensor_cpu, gt_landmarks_cpu, pred_tensor_cpu, number_list):   
    batch_size = input_tensor_cpu.shape[0]
    if batch_size > 8:
        batch_size = 8
    fig_col = int(batch_size/2)
    total_input = []
    total_gt_landmark = []
    total_pd_landmark = []
    total_pd_heatmap_l1 = []
    total_pd_heatmap_l2 = []
    for num in range(batch_size):
        input_slice = input_tensor_cpu[num][0]
        gt_landmark = gt_landmarks_cpu[num]
        pred_slices = pred_tensor_cpu[num]
        number_slice = number_list[num]

        gt_landmarks = np.array(gt_landmark, dtype=np.int64)

        pred_landmarks = []
        for i in range(pred_slices.shape[0]): ## pred_slices.shape[0] == num landmark
            landmark = convert_heatmap_to_landmark(pred_slices[i, :, :]) 
            pred_landmarks.append(landmark)
        pred_landmarks = np.array(pred_landmarks)

        total_input.append(input_slice)
        total_gt_landmark.append(gt_landmarks)
        total_pd_landmark.append(pred_landmarks)
        total_pd_heatmap_l1.append(pred_slices[0])
        total_pd_heatmap_l2.append(pred_slices[1])

    ## input + box
    fig = plt.figure() 
    gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
    for num in range(len(total_input)):
        ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
        ax.imshow(total_input[num], cmap='gray')
        gt = total_gt_landmark[num] ############# restore original scale (128 -> 512)
        pd = total_pd_landmark[num] * 4 ############# restore original scale (128 -> 512)
        ax.add_patch(patches.Rectangle(gt[0], gt[1][0] - gt[0][0], gt[1][1] - gt[0][1], edgecolor = 'red', fill=False))
        ax.add_patch(patches.Rectangle(pd[0], pd[1][0] - pd[0][0], pd[1][1] - pd[0][1], edgecolor = 'blue', fill=False))
        title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
        ax.set_title(title)
        ax.axis('off')
    plt.tight_layout()
    plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
    plt.margins(0,0)
    plt.savefig(path_save + "_Result.jpg")
    plt.close()

    ## land 1 heat map
    fig = plt.figure() 
    gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
    for num in range(len(total_input)):
        ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
        ax.imshow(total_pd_heatmap_l1[num])
        gt = total_gt_landmark[num]
        pd = total_pd_landmark[num]
        ax.scatter(int(gt[0][0]/4), int(gt[0][1]/4), c='r', s=10)
        ax.scatter(pd[0][0], pd[0][1], c='blue', s=10)
        title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
        ax.set_title(title)
        ax.axis('off')
    plt.tight_layout()
    plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
    plt.margins(0,0)
    plt.savefig(path_save + "_Land1.jpg")
    plt.close()

    ## land 2 heat map
    fig = plt.figure() 
    gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
    for num in range(len(total_input)):
        ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
        ax.imshow(total_pd_heatmap_l2[num])
        gt = total_gt_landmark[num]
        pd = total_pd_landmark[num]
        ax.scatter(int(gt[1][0]/4), int(gt[1][1]/4), c='r', s=10)
        ax.scatter(pd[1][0], pd[1][1], c='blue', s=10)
        title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
        ax.set_title(title)
        ax.axis('off')
    plt.tight_layout()
    plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
    plt.margins(0,0)
    plt.savefig(path_save + "_Land2.jpg")
    plt.close()

def save_result_image_classification(path_save, input_tensor_cpu, gt_landmarks_cpu, pred_tensor_cpu, number_list, classification_list):   
    batch_size = input_tensor_cpu.shape[0]
    if batch_size > 8:
        batch_size = 8
    fig_col = int(batch_size/2)
    total_input = []
    total_gt_landmark = []
    total_pd_landmark = []
    total_pd_heatmap_l1 = []
    total_pd_heatmap_l2 = []
    for num in range(batch_size):
        input_slice = input_tensor_cpu[num][0]
        gt_landmark = gt_landmarks_cpu[num]
        pred_slices = pred_tensor_cpu[num]
        number_slice = number_list[num]

        gt_landmarks = np.array(gt_landmark, dtype=np.int64)

        pred_landmarks = []
        for i in range(pred_slices.shape[0]): ## pred_slices.shape[0] == num landmark
            landmark = convert_heatmap_to_landmark(pred_slices[i, :, :]) 
            pred_landmarks.append(landmark)
        pred_landmarks = np.array(pred_landmarks)

        total_input.append(input_slice)
        total_gt_landmark.append(gt_landmarks)
        total_pd_landmark.append(pred_landmarks)
        total_pd_heatmap_l1.append(pred_slices[0])
        total_pd_heatmap_l2.append(pred_slices[1])

    ## input + box
    fig = plt.figure() 
    gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
    for num in range(len(total_input)):
        ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
        ax.imshow(total_input[num], cmap='gray')
        gt = total_gt_landmark[num] ############# restore original scale (128 -> 512)
        pd = total_pd_landmark[num] * 4 ############# restore original scale (128 -> 512)
        ax.add_patch(patches.Rectangle(gt[0], gt[1][0] - gt[0][0], gt[1][1] - gt[0][1], edgecolor = 'red', fill=False))
        ax.add_patch(patches.Rectangle(pd[0], pd[1][0] - pd[0][0], pd[1][1] - pd[0][1], edgecolor = 'blue', fill=False))
        title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
        ax.set_title(title)
        ax.axis('off')
    plt.tight_layout()
    plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
    plt.margins(0,0)
    plt.savefig(path_save + "_Result.jpg")
    plt.close()

    ## land 1 heat map
    fig = plt.figure() 
    gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
    for num in range(len(total_input)):
        ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
        ax.imshow(total_pd_heatmap_l1[num])
        gt = total_gt_landmark[num]
        pd = total_pd_landmark[num]
        ax.scatter(int(gt[0][0]/4), int(gt[0][1]/4), c='r', s=10)
        ax.scatter(pd[0][0], pd[0][1], c='blue', s=10)
        title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
        ax.set_title(title)
        ax.axis('off')
    plt.tight_layout()
    plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
    plt.margins(0,0)
    plt.savefig(path_save + "_Land1.jpg")
    plt.close()

    ## land 2 heat map
    fig = plt.figure() 
    gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
    for num in range(len(total_input)):
        ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
        ax.imshow(total_pd_heatmap_l2[num])
        gt = total_gt_landmark[num]
        pd = total_pd_landmark[num]
        ax.scatter(int(gt[1][0]/4), int(gt[1][1]/4), c='r', s=10)
        ax.scatter(pd[1][0], pd[1][1], c='blue', s=10)
        title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
        if classification_list[num] == 1:
            title += "_T"
        else:
            title += "_Z"
        ax.set_title(title)
        ax.axis('off')
    plt.tight_layout()
    plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
    plt.margins(0,0)
    plt.savefig(path_save + "_Land2.jpg")
    plt.close()

if __name__ == "__main__": 
    ### slice 숫자 0000 형태로 바꾸는 코드 필요
    dir_data = "Z:/Backup/Users/kys/BoundingBox/data/raw/Colorectal/2010/"
    dir_result = "Z:/Backup/Users/kys/BoundingBox/data/processed/heatmap_128_LN2_2010_HU/"
    # prepare_data(dir_data, dir_result, sigma=10)
    # prepare_data_3slice(dir_data, dir_result, sigma=10)
    # dir_data = "Z:/Backup/Users/kys/BoundingBox/data/val_rest/"
    # dir_result = "Z:/Backup/Users/kys/BoundingBox/data/processed/Validation_rest/"
    # prepare_data_val(dir_data, dir_result, sigma=10, num_landmark=2)
  
    ## pre 이후에
    dir_data = "Z:/Backup/Users/kys/BoundingBox/data/processed/heatmap_128_LN2_2010_HU/"
    # count_patient_num(dir_data + "train/")
    divide_data_patient(dir_data, 0.1, 0)
    type_dir = ["train/", "test/"]
    for t in type_dir:
        divide_balance(dir_data + t)
    # divide_balance("Z:/Backup/Users/kys/BoundingBox/data/processed/Validation_rest/val/")

    ori_path = "Z:/Backup/Users/kys/BoundingBox/data/processed/heatmap_128_LN2_2012_HU/train/"
    # move_path = "Z:/Backup/Users/kys/BoundingBox/data/processed/heatmap_128_LN2_2011_HU/train_Zero/"
    # move_file(ori_path, move_path)
    copy_path = "Z:/Backup/Users/kys/BoundingBox/data/processed/heatmap_128_LN2_20All3_HU/train/"
    

    root_dir = "Z:/Backup/Users/kys/BoundingBox/data/processed/"
    year_dir = ["LN2_2011_HU_3slice/", "LN2_2012_HU_3slice/", "LN2_2014_HU_3slice/"]
    year_dir = ["heatmap_128_LN2_2010_HU/"]
    type_dir = ["train/", "test/"]
    for year in year_dir:
        copy_file(root_dir + year + "train/", copy_path)
    # dir_data = "Z:/Backup/Users/kys/BoundingBox/data/"
    # GetDataInformationFromDicom(dir_data)

    # dir_data = "Z:/Backup/Users/kys/BoundingBox/data/processed/Validation/all/"
    # rename_slice(dir_data)
    # dir_data = "Z:/Backup/Users/kys/BoundingBox/data/processed/LN2_2014_HU_3slice/"
    # count_patient_num(dir_data)

