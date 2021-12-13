import os
import datetime
import logging
from caffe2.python.workspace import GpuDeviceType

import matplotlib.pyplot as plt
import numpy as np
import time
import openpyxl

import torch
import torch.optim as optim    
from torch.optim import lr_scheduler
import torch.nn as nn  
from torch.utils.tensorboard import SummaryWriter
from torch.utils.data import DataLoader

from network import Single_hourglass_network
from network2 import Single_hourglass_network as Net2

from BoundingBoxDataSet import BoundingBoxDataSet
import score
import data

# import gradcam as cam
from gradcam import gradcam
from gradcam.util.image import show_cam_on_image, \
                                         deprocess_image, \
                                         preprocess_image

import random
import cv2
import pandas as pd
os.environ['KMP_DUPLICATE_LIB_OK']='True'
#########################################################################################################
## tensorboard --logdir=Z:/Backup/Users/kys/BoundingBox/result/210810_17h46m_LN2_All_MSE_S4_Com1/Tensorboard/
########## Args ##########
def train():
    GPU_Name = "cuda:0"
    device = torch.device(GPU_Name if torch.cuda.is_available() else "cpu")
    gpu_info = torch.cuda.get_device_name(device)
    
    DIR_RESULT_PARENT = "Z:/Backup/Users/kys/BoundingBox/result/"

    DIR_DATA = "Z:/Backup/Users/kys/BoundingBox/data/processed/heatmap_128_LN2_20All3_HU/"
    DIR_DATA_TRAIN = DIR_DATA + "train/" 
    DIR_DATA_TEST = DIR_DATA + "test/" 
    PATH_PRETRAINED = None

    # check_path = "210912_16h16m_LN2_All3_S4_M20_CW005_AddClassify_Com1/"
    # Check_last = [x for x in os.listdir(f"{DIR_RESULT_PARENT}{check_path}Checkpoint/") if not "Best" in x][-1]
    # PATH_PRETRAINED = f"{DIR_RESULT_PARENT}{check_path}Checkpoint/{Check_last}"
    # PATH_PRETRAINED_2 = f"{DIR_RESULT_PARENT}{check_path}Checkpoint_2/{Check_last}"

    COMPUTER = ""
    if "2080" in gpu_info:
        COMPUTER = "Com1"
    elif "3060" in gpu_info:
        COMPUTER = "Com2"
    else:
        print(f"Check Gpu info - {gpu_info}")
    NUM_LANDMARKS = 2
    LR = 1e-4/2
    Batch_Size = 8
    NUM_EPOCH = 20
    HM_Size = 4
    Loss_Weight = None
    Loss_Weight = "1.5"
    Best_Dist = 110.0

    network = Net2(num_landmarks = NUM_LANDMARKS)

    opti = optim.Adam(network.parameters(), lr=LR)
    # loss_function = nn.BCEWithLogitsLoss()
    loss_function = nn.MSELoss()
    # scheduler = lr_scheduler.ReduceLROnPlateau(opti, 'min', patience=2)
    # SCHEDULER = lr_scheduler.ReduceLROnPlateau(OPTI, 'min' if OUT_DIM > 1 else 'max', patience=2)

    NOTE_NETWORK = f"Single_hourglass_network(num_landmarks = {NUM_LANDMARKS})"
    NOTE_OPTI = f"optim.Adam(network.parameters(), lr={LR})"
    # NOTE_LOSS = "nn.BCEWithLogitsLoss()"
    NOTE_LOSS = "nn.MSELoss()"
    # NOTE_SCHEDULER = "lr_scheduler.ReduceLROnPlateau(opti, 'min', patience=2)"

    ## Name Result
    dir_data_split = DIR_DATA.split('_')
    Data_Year = ''
    for year in dir_data_split:
        if "20" in year:
            Data_Year = year
    Data_Year = Data_Year[2:]
    Loss_F = NOTE_LOSS.split(".")[1][:3]
    lw_val = ''
    if Loss_Weight is not None:
        lw_val = f"_M{Loss_Weight.split('.')[0]}{Loss_Weight.split('.')[1]}"

    NAME_RESULT = f"LN{NUM_LANDMARKS}_{Data_Year}_S{HM_Size}{lw_val}"
    
    ## grad cam
    #########################################################################################################
    # args = gradcam.get_args()
    # cam = gradcam.CAM(model=network, target_layer=network.hourglass_block.conv_layer, use_cuda=args.use_cuda)
    # target_layer_names=["hourglass_block.conv_final.weight"],

    ########## Init ##########
    # directories
    today = str(datetime.date.today().strftime('%y%m%d'))
    now = str(datetime.datetime.now().strftime('%Hh%Mm'))

    DIR_RESULT = f"{DIR_RESULT_PARENT}{today}_{now}_{NAME_RESULT}_{COMPUTER}/"
    DIR_TRAIN_OUTPUT = f"{DIR_RESULT}IMG_Train/"
    DIR_TEST_OUTPUT = f"{DIR_RESULT}IMG_Test/"
    DIR_TENSORBOARD = f"{DIR_RESULT}Tensorborad/"
    DIR_CHECKPOINT = f"{DIR_RESULT}Checkpoint/"
    if PATH_PRETRAINED is not None:   
        print("PRETERAIN")
        DIR_RESULT = f"{PATH_PRETRAINED[:PATH_PRETRAINED.find('/Checkpoint')]}/" 
        DIR_TRAIN_OUTPUT = f"{DIR_RESULT}IMG_Train/"
        DIR_TEST_OUTPUT = f"{DIR_RESULT}IMG_Test/"
        DIR_TENSORBOARD = f"{DIR_RESULT}Tensorborad/"
        DIR_CHECKPOINT = f"{DIR_RESULT}Checkpoint/"
    os.makedirs(DIR_RESULT, exist_ok=True)
    os.makedirs(DIR_TRAIN_OUTPUT, exist_ok=True)
    os.makedirs(DIR_TEST_OUTPUT, exist_ok=True)
    os.makedirs(DIR_TENSORBOARD, exist_ok=True)
    os.makedirs(DIR_CHECKPOINT, exist_ok=True)
    print(f"{today}_{now}_{NAME_RESULT}_{COMPUTER} Train START")

    # device
    device = torch.device(GPU_Name if torch.cuda.is_available() else "cpu")
    torch.cuda.set_device(device)
    print(f"Using device = {device} - {torch.cuda.get_device_name(device)}")
    network.to(device=device)

    # log
    path_log = f"{DIR_RESULT}/Log.txt"
    logging.basicConfig(filename=path_log, level=logging.INFO, format='%(levelname)s : %(message)s')
    logging.info("DIR_DATA_TRAIN: " + DIR_DATA_TRAIN)
    logging.info("DIR_DATA_TEST: " + DIR_DATA_TEST)
    logging.info(f"Using device = {device} - {torch.cuda.get_device_name(device)}")
    logging.info("NUM_EPOCH: " + str(NUM_EPOCH))
    logging.info("LR: " + str(LR))
    logging.info("Batch_Size: " + str(Batch_Size))
    logging.info("HM Size: " + str(HM_Size))
    logging.info("Network: " + NOTE_NETWORK)
    logging.info("N_parameters: " + str(sum(p.numel() for p in network.parameters() if p.requires_grad)))
    logging.info("Opti: " + NOTE_OPTI)
    logging.info("Loss_function: " + NOTE_LOSS)
    # logging.info("scheduler: " + NOTE_SCHEDULER)

    # writer for tensorboard
    summary_writer = SummaryWriter(DIR_TENSORBOARD)

    # Init result excel file
    PATH_EXCEL_RESULTS = f"{DIR_RESULT}/{NAME_RESULT}_Result.xlsx"
    if os.path.isfile(PATH_EXCEL_RESULTS) is False:
        wb = openpyxl.Workbook()
        worksheet = wb.active
        worksheet.append(["Epoch", "Train_Loss", "Train_Dist", "Train_Time", "Test_Loss", "Test_Dist", "Test_Time"])
        wb.save(PATH_EXCEL_RESULTS)

    test_dataset = BoundingBoxDataSet(DIR_DATA_TEST, flag_shuffle=False, sigma=HM_Size)
    test_loader = DataLoader(test_dataset, batch_size=Batch_Size, shuffle=False, num_workers=0)
    del test_dataset
    # load pretrained model
    NUM_EPOCH += 1
    epoch_start = 1
    train_idx_step = 0
    test_idx_step = 0
    train_loss = []
    test_loss = []
    train_dist = []
    test_dist = []
    if PATH_PRETRAINED is not None:
        ## pre-train model load
        ckpt = torch.load(PATH_PRETRAINED, map_location=device)
        ckpt_epoch = ckpt['epoch']
        network.load_state_dict(ckpt['model_state_dict'])
        epoch_start = ckpt_epoch + 1
        train_idx_step = ckpt['train_idx_step']
        test_idx_step = ckpt['test_idx_step']

        print(f"{ckpt_epoch} Epoch Model loaded")
        logging.info(f'Model loaded, epoch={ckpt_epoch}, idx_step={train_idx_step}, test_idx_step={test_idx_step}')
        del ckpt
        ## pre-train result load
        pre_results = pd.read_excel(f"{DIR_RESULT}{NAME_RESULT}_Result.xlsx")
        for num in range(len(pre_results)):
            train_loss.append(pre_results['Train_Loss'][num])
            test_loss.append(pre_results['Test_Loss'][num])
            train_dist.append(pre_results['Train_Dist'][num])
            test_dist.append(pre_results['Test_Dist'][num])

    ########## iterate for epoch times ##########
    for epoch in range(epoch_start, NUM_EPOCH):
        ########## Train ##########
        network.train()
        time_start = time.time()

        # Load data
        train_dataset = BoundingBoxDataSet(DIR_DATA_TRAIN, flag_shuffle=True, sigma=HM_Size)
        train_loader = DataLoader(train_dataset, batch_size=Batch_Size, shuffle=True, num_workers=0)
        del train_dataset
    
        result = learning(train_loader, network, device, opti, loss_function, summary_writer, DIR_TRAIN_OUTPUT, epoch, Loss_Weight, idx_step=train_idx_step)
        del train_loader

        train_loss_epoch = result[0]
        train_dist_epoch = result[1]
        train_idx_step = result[2]

        time_end = time.time()
        train_time_epoch = (time_end - time_start) / 60

        # train results
        print(f'{epoch} Epoch training finished. Loss={train_loss_epoch}, Dist={train_dist_epoch}, Time(m)={round(train_time_epoch, 2)}')                
        logging.info(f'{epoch} Epoch training finished. Loss={train_loss_epoch}, Dist={train_dist_epoch}, Time(m)={train_time_epoch}')
        summary_writer.add_scalar('Loss/train', train_loss_epoch, epoch)
        summary_writer.add_scalar('Dist/train', train_dist_epoch, epoch)

        ########## test set ##########
        network.eval()
        time_start = time.time()
        with torch.no_grad():
            result = learning(test_loader, network, device, None, loss_function, summary_writer, DIR_TEST_OUTPUT, epoch, Loss_Weight, idx_step=test_idx_step)
        
        test_loss_epoch = result[0]
        test_dist_epoch = result[1]
        test_idx_step = result[2]

        time_end = time.time()
        test_time_epoch = (time_end - time_start) / 60
 
        # test results
        summary_writer.add_scalar('Loss/test', test_loss_epoch, epoch)
        summary_writer.add_scalar('Dist/test', test_dist_epoch, epoch)

        print(f'{epoch} Epoch test finished. Loss={test_loss_epoch}, Dist={test_dist_epoch}, Time(m)={round(test_time_epoch, 2)}')      
        logging.info(f'{epoch} Epoch test finished. Loss={test_loss_epoch}, Dist={test_dist_epoch}, Time(m)={test_time_epoch}')
        
        epoch_time = (train_time_epoch + test_time_epoch)
        Rest_Time = epoch_time * (NUM_EPOCH - epoch - 1)
        Rest_H = int(Rest_Time // 60)
        Rest_M = Rest_Time % 60
        print(f"{epoch} Epoch Finish Total Time(m) = {round(epoch_time,2)}, Rest Time = About {Rest_H} h {round(Rest_M,2)} m")

        ########## Save oeverall results ##########
        # save model     
        str_num = len(str(epoch))
        zero_need = 4 - str_num
        z = ''
        for i in range(zero_need):
            z += '0'
        new_epoch = f"{z}{epoch}"
        path_model = f"{DIR_CHECKPOINT}Check_epoch_{new_epoch}.pth"
        torch.save({'epoch':epoch, 'train_idx_step':train_idx_step, 'test_idx_step':test_idx_step, 'model_state_dict':network.state_dict()}, path_model)
        if train_dist_epoch < Best_Dist:
            Best_Dist = train_dist_epoch
            path_model = f"{DIR_CHECKPOINT}Check_epoch_Best.pth"
            torch.save({'epoch':epoch, 'train_idx_step':train_idx_step, 'test_idx_step':test_idx_step, 'model_state_dict':network.state_dict()}, path_model)
            print(f"Current Best model is {epoch} epoch model")

        wb = openpyxl.load_workbook(PATH_EXCEL_RESULTS)
        ws = wb.active
        result_list = [epoch, train_loss_epoch, train_dist_epoch, train_time_epoch, test_loss_epoch, test_dist_epoch, test_time_epoch]
        ws.append(result_list)
        wb.save(PATH_EXCEL_RESULTS)

        train_loss.append(train_loss_epoch)
        test_loss.append(test_loss_epoch)
        train_dist.append(train_dist_epoch)
        test_dist.append(test_dist_epoch)

    print(f"{NAME_RESULT} FINISH")
    epoch_x = np.array(range(len(train_loss)))
    epoch_x += 1
    plt.plot(epoch_x, train_loss, 'g', label='Train Loss')
    plt.plot(epoch_x, test_loss, 'r', label='Test Loss')
    plt.title('Train and Test Loss')
    plt.xlabel("Epoch")
    plt.ylabel("Loss")
    plt.legend(loc='upper right')
    plt.tight_layout()
    plt.savefig(f"{DIR_RESULT}Result_Loss.jpg")
    plt.close()

    plt.plot(epoch_x, train_dist, 'g', label='Train Dist')
    plt.plot(epoch_x, test_dist, 'r', label='Test Dist')
    plt.title('Train and Test Dist')
    plt.xlabel("Epoch")
    plt.ylabel("Dist")
    plt.legend(loc='upper right')
    plt.tight_layout()
    plt.savefig(f"{DIR_RESULT}Result_Dist.jpg")
    plt.close()

def train_3slice_2stg():
    DIR_RESULT_PARENT = "Z:/Backup/Users/kys/BoundingBox/result/"

    DIR_DATA = "Z:/Backup/Users/kys/BoundingBox/data/processed/LN2_20All_HU_3slice/"
    # DIR_DATA = "Z:/Backup/Users/kys/BoundingBox/data/processed/3lice_Test/"
    DIR_DATA_TRAIN = DIR_DATA + "train/" 
    DIR_DATA_TEST = DIR_DATA + "test/"  
    PATH_PRETRAINED = None

     # check_path = "210912_16h16m_LN2_All3_S4_M20_CW005_AddClassify_Com1/"
    # Check_last = os.listdir(f"{DIR_RESULT_PARENT}{check_path}Checkpoint/")[-1]
    # PATH_PRETRAINED = f"{DIR_RESULT_PARENT}{check_path}Checkpoint/{Check_last}"
    # PATH_PRETRAINED_2 = f"{DIR_RESULT_PARENT}{check_path}Checkpoint_2/{Check_last}"
    # print(f"Pretrain model = {PATH_PRETRAINED}")

    COMPUTER = "Com1"
    NUM_LANDMARKS = 2
    LR = 1e-4/2
    Batch_Size = 8
    NUM_EPOCH = 20
    HM_Size = 4
    Loss_Weight = None
    # Loss_Weight = "1.5"

    network = Net2(input_ch = 3, num_landmarks = NUM_LANDMARKS)
    network_2 = Net2(input_ch = 3, num_landmarks = NUM_LANDMARKS)
    # network = Single_hourglass_network(num_landmarks = NUM_LANDMARKS, hg_depth=3)
    opti = optim.Adam(network.parameters(), lr=LR)
    # loss_function = nn.BCEWithLogitsLoss()
    loss_function = nn.MSELoss()
    # scheduler = lr_scheduler.ReduceLROnPlateau(opti, 'min', patience=2)
    # SCHEDULER = lr_scheduler.ReduceLROnPlateau(OPTI, 'min' if OUT_DIM > 1 else 'max', patience=2)

    NOTE_NETWORK = f"Single_hourglass_network(num_landmarks = {NUM_LANDMARKS})"
    NOTE_OPTI = f"optim.Adam(network.parameters(), lr={LR})"
    # NOTE_LOSS = "nn.BCEWithLogitsLoss()"
    NOTE_LOSS = "nn.MSELoss()"
    # NOTE_SCHEDULER = "lr_scheduler.ReduceLROnPlateau(opti, 'min', patience=2)"

    ## Name Result
    dir_data_split = DIR_DATA.split('_')
    Data_Year = ''
    for year in dir_data_split:
        if "20" in year:
            Data_Year = year
    Data_Year = Data_Year[2:]
    Loss_F = NOTE_LOSS.split(".")[1][:3]
    lw_val = ''
    if Loss_Weight is not None:
        lw_val = f"_M{Loss_Weight.split('.')[0]}{Loss_Weight.split('.')[1]}"
        Batch_Size = int(Batch_Size/2)
        
    NAME_RESULT = f"LN{NUM_LANDMARKS}_{Data_Year}_S{HM_Size}{lw_val}_3Slice_2stg"
    
    ## grad cam
    #########################################################################################################
    # args = gradcam.get_args()
    # cam = gradcam.CAM(model=network, target_layer=network.hourglass_block.conv_layer, use_cuda=args.use_cuda)
    # target_layer_names=["hourglass_block.conv_final.weight"],

    ########## Init ##########
    # directories
    today = str(datetime.date.today().strftime('%y%m%d'))
    now = str(datetime.datetime.now().strftime('%Hh%Mm'))

    DIR_RESULT = f"{DIR_RESULT_PARENT}{today}_{now}_{NAME_RESULT}_{COMPUTER}/"
    DIR_TRAIN_OUTPUT = f"{DIR_RESULT}IMG_Train/"
    DIR_TEST_OUTPUT = f"{DIR_RESULT}IMG_Test/"
    DIR_TENSORBOARD = f"{DIR_RESULT}Tensorborad/"
    DIR_CHECKPOINT = f"{DIR_RESULT}Checkpoint/"
    DIR_CHECKPOINT_2 = f"{DIR_RESULT}Checkpoint_2/"
    
    if PATH_PRETRAINED is not None:   
        DIR_RESULT = f"{PATH_PRETRAINED[:PATH_PRETRAINED.find('/Checkpoint')]}/" 
        DIR_TRAIN_OUTPUT = f"{DIR_RESULT}IMG_Train/"
        DIR_TEST_OUTPUT = f"{DIR_RESULT}IMG_Test/"
        DIR_TENSORBOARD = f"{DIR_RESULT}Tensorborad/"
        DIR_CHECKPOINT = f"{DIR_RESULT}Checkpoint/"
        DIR_CHECKPOINT_2 = f"{DIR_RESULT}Checkpoint_2/"
    
    os.makedirs(DIR_RESULT, exist_ok=True)
    os.makedirs(DIR_TRAIN_OUTPUT, exist_ok=True)
    os.makedirs(DIR_TEST_OUTPUT, exist_ok=True)
    os.makedirs(DIR_TENSORBOARD, exist_ok=True)
    os.makedirs(DIR_CHECKPOINT, exist_ok=True)
    os.makedirs(DIR_CHECKPOINT_2, exist_ok=True)
    print(f"{DIR_RESULT[:-1]} Train START")

    # log
    path_log = f"{DIR_RESULT}/Log.txt"
    logging.basicConfig(filename=path_log, level=logging.INFO, format='%(levelname)s : %(message)s')
    logging.info("DIR_DATA_TRAIN: " + DIR_DATA_TRAIN)
    logging.info("DIR_DATA_TEST: " + DIR_DATA_TEST)
    logging.info("NUM_EPOCH: " + str(NUM_EPOCH))
    logging.info("LR: " + str(LR))
    logging.info("Batch_Size: " + str(Batch_Size))
    logging.info("HM Size: " + str(HM_Size))
    # logging.info("FLAG_BALANCE: " + str(FLAG_BALANCE))
    logging.info("Network: " + NOTE_NETWORK)
    logging.info("N_parameters: " + str(sum(p.numel() for p in network.parameters() if p.requires_grad)))
    logging.info("Opti: " + NOTE_OPTI)
    logging.info("Loss_function: " + NOTE_LOSS)
    # logging.info("scheduler: " + NOTE_SCHEDULER)

    # writer for tensorboard
    summary_writer = SummaryWriter(DIR_TENSORBOARD)

    # Init result excel file
    PATH_EXCEL_RESULTS = f"{DIR_RESULT}/{NAME_RESULT}_Result.xlsx"
    if os.path.isfile(PATH_EXCEL_RESULTS) is False:
        wb = openpyxl.Workbook()
        worksheet = wb.active
        worksheet.append(["Epoch", "Train_Loss", "Train_Dist", "Train_Loss_2", "Train_Dist_2", "Train_Time", "Test_Loss", "Test_Dist", "Test_Loss_2", "Test_Dist_2", "Test_Time"])
        wb.save(PATH_EXCEL_RESULTS)

    # device
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    torch.cuda.set_device(device)
    network.to(device=device)
    network_2.to(device=device)

    test_dataset = BoundingBoxDataSet(DIR_DATA_TEST, flag_shuffle=False, sigma=HM_Size)
    test_loader = DataLoader(test_dataset, batch_size=Batch_Size, shuffle=False, num_workers=0)
    del test_dataset
    # load pretrained model
    NUM_EPOCH += 1
    epoch_start = 1
    train_idx_step = 0
    test_idx_step = 0
    train_loss = []
    test_loss = []
    train_dist = []
    test_dist = []
    train_loss_2 = []
    test_loss_2 = []
    train_dist_2 = []
    test_dist_2 = []
    if PATH_PRETRAINED is not None:
        ## pre-train model load
        ckpt = torch.load(PATH_PRETRAINED, map_location=device)
        ckpt_epoch = ckpt['epoch']
        network.load_state_dict(ckpt['model_state_dict'])
        ckpt_2 = torch.load(PATH_PRETRAINED_2, map_location=device)
        network_2.load_state_dict(ckpt['model_state_dict'])
        epoch_start = ckpt_epoch + 1
        train_idx_step = ckpt['train_idx_step']
        test_idx_step = ckpt['test_idx_step']

        print(f"{ckpt_epoch} Epoch Model loaded")
        logging.info(f'Model loaded, epoch={ckpt_epoch}, idx_step={train_idx_step}, test_idx_step={test_idx_step}')
        del ckpt, ckpt_2
        ## pre-train result load
        pre_results = pd.read_excel(f"{DIR_RESULT}{NAME_RESULT}_Result.xlsx")
        for num in range(len(pre_results)):
            train_loss.append(pre_results['Train_Loss'][num])
            test_loss.append(pre_results['Test_Loss'][num])
            train_dist.append(pre_results['Train_Dist'][num])
            test_dist.append(pre_results['Test_Dist'][num])
            train_loss_2.append(pre_results['Train_Loss_2'][num])
            test_loss_2.append(pre_results['Test_Loss_2'][num])
            train_dist_2.append(pre_results['Train_Dist_2'][num])
            test_dist_2.append(pre_results['Test_Dist_2'][num])

    ########## iterate for epoch times ##########
    for epoch in range(epoch_start, NUM_EPOCH):
        ########## Train ##########
        network.train()
        network_2.train()
        time_start = time.time()

        # Load data
        train_dataset = BoundingBoxDataSet(DIR_DATA_TRAIN, flag_shuffle=True, sigma=HM_Size)
        train_loader = DataLoader(train_dataset, batch_size=Batch_Size, shuffle=True, num_workers=0)
        del train_dataset
        # result = learning(test_loader, network, device, None, loss_function, summary_writer, DIR_TEST_OUTPUT, epoch, idx_step=test_idx_step, is_weight=Is_Weight)
        result = learning_3slice_2stg(train_loader, network, network_2, device, opti, loss_function, summary_writer, DIR_TRAIN_OUTPUT, epoch, idx_step=train_idx_step, is_weight=Is_Weight)
        del train_loader

        train_loss_epoch = result[0]
        train_dist_epoch = result[1]
        train_idx_step = result[2]
        train_loss_epoch_2 = result[3]
        train_dist_epoch_2 = result[4]

        time_end = time.time()
        train_time_epoch = (time_end - time_start) / 60

        # train results
        print(f'{epoch} Epoch training finished. Loss={train_loss_epoch}, Loss_Sec={train_loss_epoch_2}, Dist={train_dist_epoch}, Dist_Sec={train_dist_epoch_2}, Time(m)={round(train_time_epoch, 2)}')                
        logging.info(f'{epoch} Epoch training finished. Loss={train_loss_epoch}, Loss_Sec={train_loss_epoch_2}, Dist={train_dist_epoch}, Dist_Sec={train_dist_epoch_2}, Time(m)={train_time_epoch}')
        summary_writer.add_scalar('Loss/train', train_loss_epoch, epoch)
        summary_writer.add_scalar('Loss_2/train', train_loss_epoch_2, epoch)
        summary_writer.add_scalar('Dist/train', train_dist_epoch, epoch)
        summary_writer.add_scalar('Dist_2/train', train_dist_epoch_2, epoch)

        ########## test set ##########
        network.eval()
        network_2.eval()
        time_start = time.time()
        with torch.no_grad():
            result = learning_3slice_2stg(test_loader, network, network_2, device, None, loss_function, summary_writer, DIR_TEST_OUTPUT, epoch, idx_step=test_idx_step, is_weight=Is_Weight)
        
        test_loss_epoch = result[0]
        test_dist_epoch = result[1]
        test_idx_step = result[2]
        test_loss_epoch_2 = result[3]
        test_dist_epoch_2 = result[4]

        time_end = time.time()
        test_time_epoch = (time_end - time_start) / 60
 
        # test results
        summary_writer.add_scalar('Loss/test', test_loss_epoch, epoch)
        summary_writer.add_scalar('Dist/test', test_dist_epoch, epoch)
        summary_writer.add_scalar('Loss_2/test', test_loss_epoch_2, epoch)
        summary_writer.add_scalar('Dist_2/test', test_dist_epoch_2, epoch)

        print(f'{epoch} Epoch test finished. Loss={test_loss_epoch}, Loss_Sec={test_loss_epoch_2}, Dist={test_dist_epoch}, Dist_Sec={test_dist_epoch_2}, Time(m)={round(test_time_epoch, 2)}')      
        logging.info(f'{epoch} Epoch test finished. Loss={test_loss_epoch}, Loss_Sec={test_loss_epoch_2},  Dist={test_dist_epoch}, Dist_Sec={test_dist_epoch_2}, Time(m)={test_time_epoch}')
        
        epoch_time = (train_time_epoch + test_time_epoch)
        if epoch == epoch_start:
            Rest_Time = epoch_time * (NUM_EPOCH + 1 - epoch_start)
        Rest_Time -= epoch_time
        print(f"{epoch} Epoch Finish Total Time(m) = {round(epoch_time,2)}, Rest Time(m) = About {round(Rest_Time,2)}")

        ########## Save oeverall results ##########
        # save model   
        str_num = len(str(epoch))
        zero_need = 4 - str_num
        z = ''
        for i in range(zero_need):
            z += '0'
        new_epoch = f"{z}{epoch}"  
        path_model = f"{DIR_CHECKPOINT}Check_epoch_{new_epoch}.pth"
        path_model_2 = f"{DIR_CHECKPOINT_2}Check_epoch_{new_epoch}.pth"
        torch.save({'epoch':epoch, 'train_idx_step':train_idx_step, 'test_idx_step':test_idx_step, 'model_state_dict':network.state_dict()}, path_model)
        torch.save({'model_state_dict':network_2.state_dict()}, path_model_2)

        wb = openpyxl.load_workbook(PATH_EXCEL_RESULTS)
        ws = wb.active
        result_list = [epoch, train_loss_epoch, train_dist_epoch, train_loss_epoch_2, train_dist_epoch_2, train_time_epoch, test_loss_epoch_2, test_dist_epoch_2, test_time_epoch]
        ws.append(result_list)
        wb.save(PATH_EXCEL_RESULTS)

        train_loss.append(train_loss_epoch)
        test_loss.append(test_loss_epoch)
        train_dist.append(train_dist_epoch)
        test_dist.append(test_dist_epoch)
        train_loss_2.append(train_loss_epoch_2)
        test_loss_2.append(test_loss_epoch_2)
        train_dist_2.append(train_dist_epoch_2)
        test_dist_2.append(test_dist_epoch_2)

    print(f"{NAME_RESULT} FINISH")
    epoch_x = range(len(train_loss))
    plt.plot(epoch_x, train_loss, 'g', label='Train Loss')
    plt.plot(epoch_x, test_loss, 'r', label='Test Loss')
    plt.title('Train and Test Loss')
    plt.xlabel("Epoch")
    plt.ylabel("Loss")
    plt.legend(loc='upper right')
    plt.tight_layout()
    plt.savefig(f"{DIR_RESULT}Result_Loss.jpg")
    plt.close()

    plt.plot(epoch_x, train_dist, 'g', label='Train Dist')
    plt.plot(epoch_x, test_dist, 'r', label='Test Dist')
    plt.title('Train and Test Dist')
    plt.xlabel("Epoch")
    plt.ylabel("Dist")
    plt.legend(loc='upper right')
    plt.tight_layout()
    plt.savefig(f"{DIR_RESULT}Result_Dist.jpg")
    plt.close()

    plt.plot(epoch_x, train_loss_2, 'g', label='Train Loss')
    plt.plot(epoch_x, test_loss_2, 'r', label='Test Loss')
    plt.title('Train and Test Loss')
    plt.xlabel("Epoch")
    plt.ylabel("Loss")
    plt.legend(loc='upper right')
    plt.tight_layout()
    plt.savefig(f"{DIR_RESULT}Result_Loss_2.jpg")
    plt.close()

    plt.plot(epoch_x, train_dist_2, 'g', label='Train Dist')
    plt.plot(epoch_x, test_dist_2, 'r', label='Test Dist')
    plt.title('Train and Test Dist')
    plt.xlabel("Epoch")
    plt.ylabel("Dist")
    plt.legend(loc='upper right')
    plt.tight_layout()
    plt.savefig(f"{DIR_RESULT}Result_Dist_2.jpg")
    plt.close()

def train_3slice_1stg():
    DIR_RESULT_PARENT = "Z:/Backup/Users/kys/BoundingBox/result/"

    DIR_DATA = "Z:/Backup/Users/kys/BoundingBox/data/processed/LN2_20All2_HU_3slice/"
    DIR_DATA_TRAIN = DIR_DATA + "train/" 
    DIR_DATA_TEST = DIR_DATA + "test/" 
    PATH_PRETRAINED = None
    
    # check_path = "210912_16h16m_LN2_All3_S4_M20_CW005_AddClassify_Com1/"
    # Check_last = os.listdir(f"{DIR_RESULT_PARENT}{check_path}Checkpoint/")[-1]
    # PATH_PRETRAINED = f"{DIR_RESULT_PARENT}{check_path}Checkpoint/{Check_last}"
    # print(f"Pretrain model = {PATH_PRETRAINED}")

    COMPUTER = "Com1"
    NUM_LANDMARKS = 2
    LR = 1e-4/2
    Batch_Size = 8
    NUM_EPOCH = 20
    HM_Size = 4
    Loss_Weight = None
    Loss_Weight = "2.0"


    network = Net2(input_ch = 3, num_landmarks = NUM_LANDMARKS)
    opti = optim.Adam(network.parameters(), lr=LR)
    # loss_function = nn.BCEWithLogitsLoss()
    loss_function = nn.MSELoss()
    # scheduler = lr_scheduler.ReduceLROnPlateau(opti, 'min', patience=2)
    # SCHEDULER = lr_scheduler.ReduceLROnPlateau(OPTI, 'min' if OUT_DIM > 1 else 'max', patience=2)

    NOTE_NETWORK = f"Single_hourglass_network(input_ch = 3, num_landmarks = {NUM_LANDMARKS})"
    NOTE_OPTI = f"optim.Adam(network.parameters(), lr={LR})"
    # NOTE_LOSS = "nn.BCEWithLogitsLoss()"
    NOTE_LOSS = "nn.MSELoss()"
    # NOTE_SCHEDULER = "lr_scheduler.ReduceLROnPlateau(opti, 'min', patience=2)"

    ## Name Result
    dir_data_split = DIR_DATA.split('_')
    Data_Year = ''
    for year in dir_data_split:
        if "20" in year:
            Data_Year = year
    Data_Year = Data_Year[2:]
    # Loss_F = NOTE_LOSS.split(".")[1][:3]
    lw_val = ''
    if Loss_Weight is not None:
        lw_val = f"_M{Loss_Weight.split('.')[0]}{Loss_Weight.split('.')[1]}"

    NAME_RESULT = f"LN{NUM_LANDMARKS}_{Data_Year}_S{HM_Size}{lw_val}_3Slice_1stg"
    
    ########## Init ##########
    # directories
    today = str(datetime.date.today().strftime('%y%m%d'))
    now = str(datetime.datetime.now().strftime('%Hh%Mm'))

    DIR_RESULT = f"{DIR_RESULT_PARENT}{today}_{now}_{NAME_RESULT}_{COMPUTER}/"
    DIR_TRAIN_OUTPUT = f"{DIR_RESULT}IMG_Train/"
    DIR_TEST_OUTPUT = f"{DIR_RESULT}IMG_Test/"
    DIR_TENSORBOARD = f"{DIR_RESULT}Tensorborad/"
    DIR_CHECKPOINT = f"{DIR_RESULT}Checkpoint/"
    if PATH_PRETRAINED is not None:   
        print("PRETERAIN")
        DIR_RESULT = f"{PATH_PRETRAINED[:PATH_PRETRAINED.find('/Checkpoint')]}/" 
        DIR_TRAIN_OUTPUT = f"{DIR_RESULT}IMG_Train/"
        DIR_TEST_OUTPUT = f"{DIR_RESULT}IMG_Test/"
        DIR_TENSORBOARD = f"{DIR_RESULT}Tensorborad/"
        DIR_CHECKPOINT = f"{DIR_RESULT}Checkpoint/"
    os.makedirs(DIR_RESULT, exist_ok=True)
    os.makedirs(DIR_TRAIN_OUTPUT, exist_ok=True)
    os.makedirs(DIR_TEST_OUTPUT, exist_ok=True)
    os.makedirs(DIR_TENSORBOARD, exist_ok=True)
    os.makedirs(DIR_CHECKPOINT, exist_ok=True)
    print(f"{today}_{now}_{NAME_RESULT}_{COMPUTER} Train START")

    # device
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    torch.cuda.set_device(device)
    print(f"Using device = {device} - {torch.cuda.get_device_name(device)}")
    network.to(device=device)

    # log
    path_log = f"{DIR_RESULT}/Log.txt"
    logging.basicConfig(filename=path_log, level=logging.INFO, format='%(levelname)s : %(message)s')
    logging.info("DIR_DATA_TRAIN: " + DIR_DATA_TRAIN)
    logging.info("DIR_DATA_TEST: " + DIR_DATA_TEST)
    logging.info(f"Using device = {device} - {torch.cuda.get_device_name(device)}")
    logging.info("NUM_EPOCH: " + str(NUM_EPOCH))
    logging.info("LR: " + str(LR))
    logging.info("Batch_Size: " + str(Batch_Size))
    logging.info("HM Size: " + str(HM_Size))
    logging.info("Network: " + NOTE_NETWORK)
    logging.info("N_parameters: " + str(sum(p.numel() for p in network.parameters() if p.requires_grad)))
    logging.info("Opti: " + NOTE_OPTI)
    logging.info("Loss_function: " + NOTE_LOSS)
    # logging.info("scheduler: " + NOTE_SCHEDULER)

    # writer for tensorboard
    summary_writer = SummaryWriter(DIR_TENSORBOARD)

    # Init result excel file
    PATH_EXCEL_RESULTS = f"{DIR_RESULT}/{NAME_RESULT}_Result.xlsx"
    if os.path.isfile(PATH_EXCEL_RESULTS) is False:
        wb = openpyxl.Workbook()
        worksheet = wb.active
        worksheet.append(["Epoch", "Train_Loss", "Train_Dist", "Train_Time", "Test_Loss", "Test_Dist", "Test_Time"])
        wb.save(PATH_EXCEL_RESULTS)

    test_dataset = BoundingBoxDataSet(DIR_DATA_TEST, flag_shuffle=False, sigma=HM_Size)
    test_loader = DataLoader(test_dataset, batch_size=Batch_Size, shuffle=False, num_workers=0)
    del test_dataset
    # load pretrained model
    NUM_EPOCH += 1
    epoch_start = 1
    train_idx_step = 0
    test_idx_step = 0
    train_loss = []
    test_loss = []
    train_dist = []
    test_dist = []
    if PATH_PRETRAINED is not None:
        ## pre-train model load
        ckpt = torch.load(PATH_PRETRAINED, map_location=device)
        ckpt_epoch = ckpt['epoch']
        network.load_state_dict(ckpt['model_state_dict'])
        epoch_start = ckpt_epoch + 1
        train_idx_step = ckpt['train_idx_step']
        test_idx_step = ckpt['test_idx_step']

        print(f"{ckpt_epoch} Epoch Model loaded")
        logging.info(f'Model loaded, epoch={ckpt_epoch}, idx_step={train_idx_step}, test_idx_step={test_idx_step}')
        del ckpt
        ## pre-train result load
        pre_results = pd.read_excel(f"{DIR_RESULT}{NAME_RESULT}_Result.xlsx")
        for num in range(len(pre_results)):
            train_loss.append(pre_results['Train_Loss'][num])
            test_loss.append(pre_results['Test_Loss'][num])
            train_dist.append(pre_results['Train_Dist'][num])
            test_dist.append(pre_results['Test_Dist'][num])

    ########## iterate for epoch times ##########
    for epoch in range(epoch_start, NUM_EPOCH):
        ########## Train ##########
        network.train()
        time_start = time.time()

        # Load data
        train_dataset = BoundingBoxDataSet(DIR_DATA_TRAIN, flag_shuffle=True, sigma=HM_Size)
        train_loader = DataLoader(train_dataset, batch_size=Batch_Size, shuffle=True, num_workers=0)
        del train_dataset
    
        # result = learning(train_loader, network, device, opti, loss_function, summary_writer, DIR_TRAIN_OUTPUT, epoch, idx_step=train_idx_step)
        result = learning_3slice_1stg(train_loader, network, device, opti, loss_function, summary_writer, DIR_TRAIN_OUTPUT, epoch, Loss_Weight, idx_step=train_idx_step)
        del train_loader

        train_loss_epoch = result[0]
        train_dist_epoch = result[1]
        train_idx_step = result[2]

        time_end = time.time()
        train_time_epoch = (time_end - time_start) / 60

        # train results
        print(f'{epoch} Epoch training finished. Loss={train_loss_epoch}, Dist={train_dist_epoch}, Time(m)={round(train_time_epoch, 2)}')                
        logging.info(f'{epoch} Epoch training finished. Loss={train_loss_epoch}, Dist={train_dist_epoch}, Time(m)={train_time_epoch}')
        summary_writer.add_scalar('Loss/train', train_loss_epoch, epoch)
        summary_writer.add_scalar('Dist/train', train_dist_epoch, epoch)

        ########## test set ##########
        network.eval()
        time_start = time.time()
        with torch.no_grad():
            # result = learning(test_loader, network, device, None, loss_function, summary_writer, DIR_TEST_OUTPUT, epoch, idx_step=test_idx_step)
            result = learning_3slice_1stg(test_loader, network, device, None, loss_function, summary_writer, DIR_TEST_OUTPUT, epoch, Loss_Weight, idx_step=test_idx_step)
        test_loss_epoch = result[0]
        test_dist_epoch = result[1]
        test_idx_step = result[2]

        time_end = time.time()
        test_time_epoch = (time_end - time_start) / 60
 
        # test results
        summary_writer.add_scalar('Loss/test', test_loss_epoch, epoch)
        summary_writer.add_scalar('Dist/test', test_dist_epoch, epoch)

        print(f'{epoch} Epoch test finished. Loss={test_loss_epoch}, Dist={test_dist_epoch}, Time(m)={round(test_time_epoch, 2)}')      
        logging.info(f'{epoch} Epoch test finished. Loss={test_loss_epoch}, Dist={test_dist_epoch}, Time(m)={test_time_epoch}')
        
        epoch_time = (train_time_epoch + test_time_epoch)
        Rest_Time = epoch_time * (NUM_EPOCH - epoch)
        Rest_H = int(Rest_Time // 60)
        Rest_M = Rest_Time % 60
        print(f"{epoch} Epoch Finish Total Time(m) = {round(epoch_time,2)}, Rest Time = About {Rest_H} h {round(Rest_M,2)} m")

        ########## Save oeverall results ##########
        # save model     
        str_num = len(str(epoch))
        zero_need = 4 - str_num
        z = ''
        for i in range(zero_need):
            z += '0'
        new_epoch = f"{z}{epoch}"
        path_model = f"{DIR_CHECKPOINT}Check_epoch_{new_epoch}.pth"
        torch.save({'epoch':epoch, 'train_idx_step':train_idx_step, 'test_idx_step':test_idx_step, 'model_state_dict':network.state_dict()}, path_model)

        wb = openpyxl.load_workbook(PATH_EXCEL_RESULTS)
        ws = wb.active
        result_list = [epoch, train_loss_epoch, train_dist_epoch, train_time_epoch, test_loss_epoch, test_dist_epoch, test_time_epoch]
        ws.append(result_list)
        wb.save(PATH_EXCEL_RESULTS)

        train_loss.append(train_loss_epoch)
        test_loss.append(test_loss_epoch)
        train_dist.append(train_dist_epoch)
        test_dist.append(test_dist_epoch)

    print(f"{NAME_RESULT} FINISH")
    epoch_x = range(len(train_loss))
    plt.plot(epoch_x, train_loss, 'g', label='Train Loss')
    plt.plot(epoch_x, test_loss, 'r', label='Test Loss')
    plt.title('Train and Test Loss')
    plt.xlabel("Epoch")
    plt.ylabel("Loss")
    plt.legend(loc='upper right')
    plt.tight_layout()
    plt.savefig(f"{DIR_RESULT}Result_Loss.jpg")
    plt.close()

    plt.plot(epoch_x, train_dist, 'g', label='Train Dist')
    plt.plot(epoch_x, test_dist, 'r', label='Test Dist')
    plt.title('Train and Test Dist')
    plt.xlabel("Epoch")
    plt.ylabel("Dist")
    plt.legend(loc='upper right')
    plt.tight_layout()
    plt.savefig(f"{DIR_RESULT}Result_Dist.jpg")
    plt.close()

def train_add_classifier():
    GPU_Name = "cuda:0"
    device = torch.device(GPU_Name if torch.cuda.is_available() else "cpu")
    gpu_info = torch.cuda.get_device_name(device)
    DIR_RESULT_PARENT = "Z:/Backup/Users/kys/BoundingBox/result/"
    DIR_DATA = "Z:/Backup/Users/kys/BoundingBox/data/processed/heatmap_128_LN2_20All4_HU/"
    DIR_DATA_TRAIN = DIR_DATA + "train/" 
    DIR_DATA_TEST = DIR_DATA + "test/" 
    PATH_PRETRAINED = None

    # check_path = "210912_16h16m_LN2_All3_S4_M20_CW005_AddClassify_Com1/"
    # Check_last = [x for x in os.listdir(f"{DIR_RESULT_PARENT}{check_path}Checkpoint/") if not "Best" in x][-1]
    # PATH_PRETRAINED = f"{DIR_RESULT_PARENT}{check_path}Checkpoint/{Check_last}"
    # print(f"Pretrain model = {PATH_PRETRAINED}")

    COMPUTER = ""
    if "2080" in gpu_info:
        COMPUTER = "Com1"
    elif "3060" in gpu_info:
        COMPUTER = "Com2"
    else:
        print(f"Check Gpu info - {gpu_info}")
    NUM_EPOCH = 20
    LR = 1e-4/2
    Batch_Size = 8
    NUM_LANDMARKS = 2
    HM_Size = 4
    Loss_weight_c = 0.001
    Loss_Weight = None
    Loss_Weight = "1.4"
    Best_Dist = 100.0

    network = Net2(num_landmarks = NUM_LANDMARKS)
    opti = optim.Adam(network.parameters(), lr=LR)
    loss_function = nn.MSELoss()
    loss_function_c = nn.BCEWithLogitsLoss()

    NOTE_NETWORK = f"Single_hourglass_network(num_landmarks = {NUM_LANDMARKS})"
    NOTE_OPTI = f"optim.Adam(network.parameters(), lr={LR})"
    NOTE_LOSS = "Heat map = nn.MSELoss(), Classification = nn.BCEWithLogitsLoss()"
    # NOTE_SCHEDULER = "lr_scheduler.ReduceLROnPlateau(opti, 'min', patience=2)"

    ## Name Result
    dir_data_split = DIR_DATA.split('_')
    Data_Year = ''
    for year in dir_data_split:
        if "20" in year:
            Data_Year = year
    Data_Year = Data_Year[2:]
    Loss_F = NOTE_LOSS.split(".")[1][:3]
    lw_val = ''
    if Loss_Weight is not None:
        lw_val = f"_M{Loss_Weight.split('.')[0]}{Loss_Weight.split('.')[1]}"
    CW_Val = str(Loss_weight_c).split('.')[0] + str(Loss_weight_c).split('.')[-1]
    NAME_RESULT = f"LN{NUM_LANDMARKS}_{Data_Year}_S{HM_Size}_CW{CW_Val}{lw_val}"

    ########## Init ##########
    # directories
    today = str(datetime.date.today().strftime('%y%m%d'))
    now = str(datetime.datetime.now().strftime('%Hh%Mm'))

    DIR_RESULT = f"{DIR_RESULT_PARENT}{today}_{now}_{NAME_RESULT}_{COMPUTER}/"
    DIR_TRAIN_OUTPUT = f"{DIR_RESULT}IMG_Train/"
    DIR_TEST_OUTPUT = f"{DIR_RESULT}IMG_Test/"
    DIR_TENSORBOARD = f"{DIR_RESULT}Tensorborad/"
    DIR_CHECKPOINT = f"{DIR_RESULT}Checkpoint/"
    if PATH_PRETRAINED is not None:   
        print("PRETERAIN")
        DIR_RESULT = f"{PATH_PRETRAINED[:PATH_PRETRAINED.find('/Checkpoint')]}/" 
        DIR_TRAIN_OUTPUT = f"{DIR_RESULT}IMG_Train/"
        DIR_TEST_OUTPUT = f"{DIR_RESULT}IMG_Test/"
        DIR_TENSORBOARD = f"{DIR_RESULT}Tensorborad/"
        DIR_CHECKPOINT = f"{DIR_RESULT}Checkpoint/"
    os.makedirs(DIR_RESULT, exist_ok=True)
    os.makedirs(DIR_TRAIN_OUTPUT, exist_ok=True)
    os.makedirs(DIR_TEST_OUTPUT, exist_ok=True)
    os.makedirs(DIR_TENSORBOARD, exist_ok=True)
    os.makedirs(DIR_CHECKPOINT, exist_ok=True)
    print(f"{today}_{now}_{NAME_RESULT}_{COMPUTER} Train START")

    # device
    device = torch.device(GPU_Name if torch.cuda.is_available() else "cpu")
    torch.cuda.set_device(device)
    print(f"Using device = {device} - {torch.cuda.get_device_name(device)}")
    network.to(device=device)

    # log
    path_log = f"{DIR_RESULT}Log.txt"
    logging.basicConfig(filename=path_log, level=logging.INFO, format='%(levelname)s : %(message)s')
    logging.info("DIR_DATA_TRAIN: " + DIR_DATA_TRAIN)
    logging.info("DIR_DATA_TEST: " + DIR_DATA_TEST)
    logging.info(f"Using device = {device} - {torch.cuda.get_device_name(device)}")
    logging.info("NUM_EPOCH: " + str(NUM_EPOCH))
    logging.info("LR: " + str(LR))
    logging.info("Batch_Size: " + str(Batch_Size))
    logging.info("HM Size: " + str(HM_Size))
    logging.info("Network: " + NOTE_NETWORK)
    logging.info("N_parameters: " + str(sum(p.numel() for p in network.parameters() if p.requires_grad)))
    logging.info("Opti: " + NOTE_OPTI)
    logging.info("Loss_function: " + NOTE_LOSS)
    logging.info(f"Classficaion Loss Weight: {Loss_weight_c}")
    # logging.info("scheduler: " + NOTE_SCHEDULER)

    # writer for tensorboard
    summary_writer = SummaryWriter(DIR_TENSORBOARD)

    # Init result excel file
    PATH_EXCEL_RESULTS = f"{DIR_RESULT}/{NAME_RESULT}_Result.xlsx"
    if os.path.isfile(PATH_EXCEL_RESULTS) is False:
        wb = openpyxl.Workbook()
        worksheet = wb.active
        worksheet.append(["Epoch", "Train_Loss", "Train_Hm_Loss", "Train_Cf_Loss", "Train_Dist", "Train_Acc", "Train_T_Acc", "Train_Z_Acc", "Train_Time", 
                                    "Test_Loss", "Test_Hm_Loss", "Test_Cf_Loss", "Test_Dist","Test_Acc", "Test_T_Acc", "Test_Z_Acc", "Test_Time"])
        wb.save(PATH_EXCEL_RESULTS)

    test_dataset = BoundingBoxDataSet(DIR_DATA_TEST, flag_shuffle=False, sigma=HM_Size)
    test_loader = DataLoader(test_dataset, batch_size=Batch_Size, shuffle=False, num_workers=0)
    del test_dataset
    # load pretrained model
    NUM_EPOCH += 1
    epoch_start = 1
    train_idx_step = 0
    test_idx_step = 0
    train_loss = []
    test_loss = []
    train_dist = []
    test_dist = []
    train_acc = []
    train_t_acc = []
    train_z_acc = []
    test_acc = []
    test_t_acc = []
    test_z_acc = []
    train_hm_loss = []
    test_hm_loss = []
    train_cf_loss = []
    test_cf_loss = []
    if PATH_PRETRAINED is not None:
        ## pre-train model load
        ckpt = torch.load(PATH_PRETRAINED, map_location=device)
        ckpt_epoch = ckpt['epoch']
        network.load_state_dict(ckpt['model_state_dict'])
        epoch_start = ckpt_epoch + 1
        train_idx_step = ckpt['train_idx_step']
        test_idx_step = ckpt['test_idx_step']

        print(f"{ckpt_epoch} Epoch Model loaded")
        logging.info(f'Model loaded, epoch={ckpt_epoch}, train_idx_step={train_idx_step}, test_idx_step={test_idx_step}')
        del ckpt
        ## pre-train result load
        pre_results = pd.read_excel(f"{DIR_RESULT}{NAME_RESULT}_Result.xlsx")
        for num in range(len(pre_results)):
            train_loss.append(pre_results['Train_Loss'][num])
            test_loss.append(pre_results['Test_Loss'][num])
            train_dist.append(pre_results['Train_Dist'][num])
            test_dist.append(pre_results['Test_Dist'][num])
            train_acc.append(pre_results['Train_Acc'][num])
            train_t_acc.append(pre_results['Train_T_Acc'][num])
            train_z_acc.append(pre_results['Train_Z_Acc'][num])
            test_acc.append(pre_results['Test_Acc'][num])
            test_t_acc.append(pre_results['Test_T_Acc'][num])
            test_z_acc.append(pre_results['Test_Z_Acc'][num])
            train_hm_loss.append(pre_results['Train_Hm_Loss'][num])
            test_hm_loss.append(pre_results['Test_Hm_Loss'][num])
            train_cf_loss.append(pre_results['Train_Cf_Loss'][num])
            test_cf_loss.append(pre_results['Test_Cf_Loss'][num])

    ########## iterate for epoch times ##########
    for epoch in range(epoch_start, NUM_EPOCH):
        ########## Train ##########
        network.train()
        time_start = time.time()

        # Load data
        train_dataset = BoundingBoxDataSet(DIR_DATA_TRAIN, flag_shuffle=True, sigma=HM_Size)
        train_loader = DataLoader(train_dataset, batch_size=Batch_Size, shuffle=True, num_workers=0)
        del train_dataset
    
        result = learning_add_classify(train_loader, network, device, opti, loss_function, loss_function_c, Loss_weight_c, summary_writer, DIR_TRAIN_OUTPUT, epoch, Loss_Weight, idx_step=train_idx_step)
        del train_loader

        train_loss_epoch = result[0]
        train_loss_hm_epoch = result[1]
        train_loss_cf_epoch = result[2]
        train_dist_epoch = result[3]
        train_idx_step = result[4]
        train_acc_epoch = result[5]
        train_acc_t_epoch = result[6]
        train_acc_z_epoch = result[7]

        time_end = time.time()
        train_time_epoch = (time_end - time_start) / 60

        # train results
        print(f'{epoch} Epoch training finished. Loss={train_loss_epoch}, Dist={train_dist_epoch}, Acc = {round(train_acc_epoch,4)*100}%, Time(m)={round(train_time_epoch, 2)}')                
        logging.info(f'{epoch} Epoch training finished. Loss={train_loss_epoch}, Dist={train_dist_epoch}, Acc = {train_acc_epoch}, Time(m)={train_time_epoch}')
        summary_writer.add_scalar('Loss/train', train_loss_epoch, epoch)
        summary_writer.add_scalar('Dist/train', train_dist_epoch, epoch)
        summary_writer.add_scalar('Acc/train', train_acc_epoch, epoch)

        ########## test set ##########
        network.eval()
        time_start = time.time()
        with torch.no_grad():
            result = learning_add_classify(test_loader, network, device, None, loss_function, loss_function_c, Loss_weight_c, summary_writer, DIR_TEST_OUTPUT, epoch, Loss_Weight, idx_step=test_idx_step)
        
        test_loss_epoch = result[0]
        test_loss_hm_epoch = result[1]
        test_loss_cf_epoch = result[2]
        test_dist_epoch = result[3]
        test_idx_step = result[4]
        test_acc_epoch = result[5]
        test_acc_t_epoch = result[6]
        test_acc_z_epoch = result[7]

        time_end = time.time()
        test_time_epoch = (time_end - time_start) / 60
 
        # test results
        summary_writer.add_scalar('Loss/test', test_loss_epoch, epoch)
        summary_writer.add_scalar('Dist/test', test_dist_epoch, epoch)
        summary_writer.add_scalar('Acc/test', test_acc_epoch, epoch)

        print(f'{epoch} Epoch test finished. Loss={test_loss_epoch}, Dist={test_dist_epoch}, Acc = {round(test_acc_epoch,4)*100}%, Time(m)={round(test_time_epoch, 2)}')      
        logging.info(f'{epoch} Epoch test finished. Loss={test_loss_epoch}, Dist={test_dist_epoch}, Acc = {train_acc_epoch}, Time(m)={test_time_epoch}')
        
        epoch_time = (train_time_epoch + test_time_epoch)
        
        Rest_Time = epoch_time * (NUM_EPOCH - epoch - 1)
        Rest_H = int(Rest_Time // 60)
        Rest_M = Rest_Time % 60
        print(f"{epoch} Epoch Finish Total Time(m) = {round(epoch_time,2)}, Rest Time = About {Rest_H} h {round(Rest_M,2)} m")

        ########## Save oeverall results ##########
        # save model     
        str_num = len(str(epoch))
        zero_need = 4 - str_num
        z = ''
        for i in range(zero_need):
            z += '0'
        new_epoch = f"{z}{epoch}"
        path_model = f"{DIR_CHECKPOINT}Check_epoch_{new_epoch}.pth"
        torch.save({'epoch':epoch, 'train_idx_step':train_idx_step, 'test_idx_step':test_idx_step, 'model_state_dict':network.state_dict()}, path_model)

        if train_dist_epoch < Best_Dist:
            Best_Dist = train_dist_epoch
            path_model = f"{DIR_CHECKPOINT}Check_epoch_Best.pth"
            torch.save({'epoch':epoch, 'train_idx_step':train_idx_step, 'test_idx_step':test_idx_step, 'model_state_dict':network.state_dict()}, path_model)
            print(f"Current Best model is {epoch} epoch model")

        wb = openpyxl.load_workbook(PATH_EXCEL_RESULTS)
        ws = wb.active
        result_list = [epoch, train_loss_epoch, train_loss_hm_epoch, train_loss_cf_epoch, train_dist_epoch, train_acc_epoch, train_acc_t_epoch, train_acc_z_epoch, train_time_epoch, 
                        test_loss_epoch, test_loss_hm_epoch, test_loss_cf_epoch, test_dist_epoch, test_acc_epoch, test_acc_t_epoch, test_acc_z_epoch, test_time_epoch]
        ws.append(result_list)
        wb.save(PATH_EXCEL_RESULTS)

        train_loss.append(train_loss_epoch)
        test_loss.append(test_loss_epoch)
        train_hm_loss.append(train_loss_hm_epoch)
        test_hm_loss.append(test_loss_hm_epoch)
        train_cf_loss.append(train_loss_cf_epoch)
        test_cf_loss.append(test_loss_cf_epoch)
        train_dist.append(train_dist_epoch)
        test_dist.append(test_dist_epoch)
        train_acc.append(train_acc_epoch)
        train_t_acc.append(train_acc_t_epoch)
        train_z_acc.append(train_acc_z_epoch)
        test_acc.append(test_acc_epoch)
        test_t_acc.append(test_acc_t_epoch)
        test_z_acc.append(test_acc_z_epoch)


    print(f"{NAME_RESULT} FINISH")
    epoch_x = np.array(range(len(train_loss)))
    epoch_x += 1
    plt.plot(epoch_x, train_loss, 'g', label='Train Loss')
    plt.plot(epoch_x, test_loss, 'r', label='Test Loss')
    plt.title('Train and Test Loss')
    plt.xlabel("Epoch")
    plt.ylabel("Loss")
    plt.legend(loc='upper right')
    plt.tight_layout()
    plt.savefig(f"{DIR_RESULT}Result_Loss.jpg")
    plt.close()

    plt.plot(epoch_x, train_hm_loss, 'g', label='Train Hm_Loss')
    plt.plot(epoch_x, test_hm_loss, 'r', label='Test Hm_Loss')
    plt.title('Train and Test Hm_Loss')
    plt.xlabel("Epoch")
    plt.ylabel("HM Loss")
    plt.legend(loc='upper right')
    plt.tight_layout()
    plt.savefig(f"{DIR_RESULT}Result_Hm_Loss.jpg")
    plt.close()

    plt.plot(epoch_x, train_cf_loss, 'g', label='Train Cf_Loss')
    plt.plot(epoch_x, test_cf_loss, 'r', label='Test Cf_Loss')
    plt.title('Train and Test Cf_Loss')
    plt.xlabel("Epoch")
    plt.ylabel("CF Loss")
    plt.legend(loc='upper right')
    plt.tight_layout()
    plt.savefig(f"{DIR_RESULT}Result_Cf_Loss.jpg")
    plt.close()

    plt.plot(epoch_x, train_dist, 'g', label='Train Dist')
    plt.plot(epoch_x, test_dist, 'r', label='Test Dist')
    plt.title('Train and Test Dist')
    plt.xlabel("Epoch")
    plt.ylabel("Dist")
    plt.legend(loc='upper right')
    plt.tight_layout()
    plt.savefig(f"{DIR_RESULT}Result_Dist.jpg")
    plt.close()

    plt.plot(epoch_x, train_acc, 'g', label='Train Acc')
    plt.plot(epoch_x, test_acc, 'r', label='Test Acc')
    plt.title('Train and Test Acc')
    plt.xlabel("Epoch")
    plt.ylabel("Acc")
    plt.legend(loc='upper right')
    plt.tight_layout()
    plt.savefig(f"{DIR_RESULT}Result_Acc.jpg")
    plt.close()

    plt.plot(epoch_x, train_t_acc, 'g', label='Train T_Acc')
    plt.plot(epoch_x, test_t_acc, 'r', label='Test T_Acc')
    plt.title('Train and Test T_Acc')
    plt.xlabel("Epoch")
    plt.ylabel("T_Acc")
    plt.legend(loc='upper right')
    plt.tight_layout()
    plt.savefig(f"{DIR_RESULT}Result_T_Acc.jpg")
    plt.close()
    
    plt.plot(epoch_x, train_z_acc, 'g', label='Train Z_Acc')
    plt.plot(epoch_x, test_z_acc, 'r', label='Test Z_Acc')
    plt.title('Train and Test Z_Acc')
    plt.xlabel("Epoch")
    plt.ylabel("Z_Acc")
    plt.legend(loc='upper right')
    plt.tight_layout()
    plt.savefig(f"{DIR_RESULT}Result_Z_Acc.jpg")
    plt.close()

def learning(data_loader, network, device, opti, loss_func, summary, img_save_dir, epoch, loss_weight, idx_step=0):
    loss_epoch = 0
    dist_epoch = 0
    idx_batch = 0
    idx_data = 0
    for batch in data_loader:
        input_tensor = batch['ct'].to(device) # torch.Size([batch, 1, 512, 512])
        gt_tensor = batch['hms'].to(device) # torch.Size([batch, 2, 128, 128])
        gt_landmarks = batch['landmarks'].to(device) 
        number_list = batch['number']

        loss = None
        # training mode
        if opti is not None:
            opti.zero_grad()
            pred_tensor = network(input_tensor)  

            loss = loss_func(pred_tensor, gt_tensor)
            if loss_weight is not None:
                for number in number_list:
                    if "ZERO" not in number:
                        loss *= float(loss_weight)
            loss.backward()
            opti.step()
        # test mode
        else:
            pred_tensor = network(input_tensor)  
            loss = loss_func(pred_tensor, gt_tensor)
            if loss_weight is not None:
                for number in number_list:
                    if "ZERO" not in number:
                        loss *= float(loss_weight)

        idx_batch += 1
        idx_step += 1
        idx_data += input_tensor.shape[0]

        loss_epoch += loss.item()
        if opti is not None:
            summary.add_scalar('Loss/Step_train', loss.item(), idx_step)
        else:
            summary.add_scalar('Loss/Step_test', loss.item(), idx_step)
        print(f'Progress: idx_data={idx_data}/{len(data_loader.dataset)}, curr loss={round(loss.item(),4)}', end='\r')

        cur_bath_size = input_tensor.shape[0]
        input_tensor_cpu = input_tensor.cpu().numpy()
        gt_landmarks_cpu = gt_landmarks.cpu().numpy()
        pred_tensor_cpu = pred_tensor.detach().cpu().numpy()

        del input_tensor, gt_tensor, gt_landmarks, pred_tensor

        for num in range(cur_bath_size):
            gt_landmark = gt_landmarks_cpu[num]
            pred_slices = pred_tensor_cpu[num]
            landmark_num = pred_slices.shape[0]

            # calculate score
            for i in range(landmark_num):
                pred_landmark = data.convert_heatmap_to_landmark(pred_slices[i, :, :])  * int(input_tensor_cpu[num][0].shape[0] / pred_slices.shape[1]) ## 128 -> 512 size
                dist = np.linalg.norm(gt_landmark-pred_landmark)
                dist_epoch += dist
        
        if idx_data == cur_bath_size or idx_data == (cur_bath_size*2):
            title = f'/{int(cur_bath_size/idx_data)}/Epoch_{epoch}'
            os.makedirs(f'{img_save_dir}/{int(cur_bath_size/idx_data)}/', exist_ok=True)
            path_save = img_save_dir + title

            data.save_result_image(path_save, input_tensor_cpu, gt_landmarks_cpu, pred_tensor_cpu, number_list)

        torch.cuda.empty_cache()
    
    loss_epoch /= idx_batch
    dist_epoch /= idx_data * landmark_num

    return [loss_epoch, dist_epoch, idx_step]

def learning_3slice_1stg(data_loader, network, device, opti, loss_func, summary, img_save_dir, epoch, loss_weight, idx_step=0):
    loss_epoch = 0
    dist_epoch = 0
    idx_batch = 0
    idx_data = 0
    pd_slice_list = []
    pd1_hmap_list = []
    pd2_hmap_list = []
    gt_hmap_list = []
    gt_landmark_list = []
    batch_size = 0
    for batch in data_loader:
        input_tensor = batch['ct'].to(device) # torch.Size([batch, 1, 3, 512, 512])
        input_tensor = input_tensor.view(input_tensor.shape[0], input_tensor.shape[2], input_tensor.shape[3], input_tensor.shape[4]) # torch.Size([batch, 3, 512, 512])
        if idx_data == 0:
            batch_size = input_tensor.shape[0]
        gt_tensor = batch['hms'].to(device) # torch.Size([batch, 2, 128, 128])
        gt_landmarks = batch['landmarks'].to(device) 
        number_list = batch['number']

        loss = None
        # training mode
        if opti is not None:
            opti.zero_grad()
            # pred_tensor = network(input_tensor[:, :, 1, :, :])  
            pred_tensor = network(input_tensor)   # torch.Size([batch, 2, 128, 128])       
            loss = loss_func(pred_tensor, gt_tensor)
            if loss_weight is not None:
                for number in number_list:
                    if "ZERO" not in number:
                        loss *= float(loss_weight)
            loss.backward()
            opti.step()
        # test mode
        else:
            pred_tensor = network(input_tensor)  
            loss = loss_func(pred_tensor, gt_tensor)
            if loss_weight is not None:
                for number in number_list:
                    if "ZERO" not in number:
                        loss *= float(loss_weight)

        idx_batch += 1
        idx_step += 1
        idx_data += input_tensor.shape[0]

        loss_epoch += loss.item()
        if opti is not None:
            summary.add_scalar('Loss/Step_train', loss.item(), idx_step)
        else:
            summary.add_scalar('Loss/Step_test', loss.item(), idx_step)
        print(f'Progress First Stage:  idx_data={idx_data}/{len(data_loader.dataset)}, curr loss={round(loss.item(),4)}', end='\r')

        cur_bath_size = input_tensor.shape[0]
        input_tensor_cpu = input_tensor.cpu().numpy()[:, 1, :, :]
        input_tensor_cpu = np.expand_dims(input_tensor_cpu, axis=1)
        gt_landmarks_cpu = gt_landmarks.cpu().numpy()
        gt_tensor_cpu = gt_tensor.cpu().numpy()
        pred_tensor_cpu = pred_tensor.detach().cpu().numpy()

        del input_tensor, gt_tensor, gt_landmarks, pred_tensor

        for num in range(cur_bath_size):
            gt_landmark = gt_landmarks_cpu[num]
            pred_slices = pred_tensor_cpu[num]
            landmark_num = pred_slices.shape[0]
            number_slice = number_list[num]
            ## for second stage
            pd_slice_list.append(number_slice)
            pd1_hmap_list.append(pred_slices[0])
            pd2_hmap_list.append(pred_slices[1])
            gt_hmap_list.append(gt_tensor_cpu[num])
            gt_landmark_list.append(gt_landmarks_cpu[num])
            # calculate score
            for i in range(landmark_num):
                pred_landmark = data.convert_heatmap_to_landmark(pred_slices[i, :, :])  * int(input_tensor_cpu[num][0].shape[0] / pred_slices.shape[1]) ## 128 -> 512 size
                dist = np.linalg.norm(gt_landmark-pred_landmark)
                dist_epoch += dist
        
        if idx_data == cur_bath_size or idx_data == (cur_bath_size*2):
            title = f'/{int(cur_bath_size/idx_data)}/Epoch_{epoch}'
            os.makedirs(f'{img_save_dir}/{int(cur_bath_size/idx_data)}/', exist_ok=True)
            path_save = img_save_dir + title

            data.save_result_image(path_save, input_tensor_cpu, gt_landmarks_cpu, pred_tensor_cpu, number_list)
        torch.cuda.empty_cache()
    
    loss_epoch /= idx_batch
    dist_epoch /= idx_data * landmark_num

    return [loss_epoch, dist_epoch, idx_step]

def learning_3slice_2stg(data_loader, network, network_2, device, opti, loss_func, summary, img_save_dir, epoch, loss_weight, idx_step=0):
    loss_epoch = 0
    dist_epoch = 0
    idx_batch = 0
    idx_data = 0
    pd_slice_list = []
    pd1_hmap_list = []
    pd2_hmap_list = []
    gt_hmap_list = []
    gt_landmark_list = []
    batch_size = 0
    for batch in data_loader:
        input_tensor = batch['ct'].to(device) # torch.Size([batch, 1, 3, 512, 512])
        input_tensor = input_tensor.view(input_tensor.shape[0], input_tensor.shape[2], input_tensor.shape[3], input_tensor.shape[4]) # torch.Size([batch, 3, 512, 512])
        if idx_data == 0:
            batch_size = input_tensor.shape[0]
        gt_tensor = batch['hms'].to(device) # torch.Size([batch, 2, 128, 128])
        gt_landmarks = batch['landmarks'].to(device) 
        number_list = batch['number']

        loss = None
        # training mode
        if opti is not None:
            opti.zero_grad()
            # pred_tensor = network(input_tensor[:, :, 1, :, :])  
            pred_tensor = network(input_tensor)   # torch.Size([batch, 2, 128, 128])

            loss = loss_func(pred_tensor, gt_tensor)
            if loss_weight is not None:
                for number in number_list:
                    if "ZERO" not in number:
                        loss *= float(loss_weight)
            loss.backward()
            opti.step()
        # test mode
        else:
            pred_tensor = network(input_tensor)  
            loss = loss_func(pred_tensor, gt_tensor)
            if loss_weight is not None:
                for number in number_list:
                    if "ZERO" not in number:
                        loss *= float(loss_weight)

        idx_batch += 1
        idx_step += 1
        idx_data += input_tensor.shape[0]

        loss_epoch += loss.item()
        if opti is not None:
            summary.add_scalar('Loss/Step_train', loss.item(), idx_step)
        else:
            summary.add_scalar('Loss/Step_test', loss.item(), idx_step)
        print(f'Progress First Stage:  idx_data={idx_data}/{len(data_loader.dataset)}, curr loss={round(loss.item(),4)}', end='\r')

        cur_bath_size = input_tensor.shape[0]
        input_tensor_cpu = input_tensor.cpu().numpy()[:, 1, :, :]
        input_tensor_cpu = np.expand_dims(input_tensor_cpu, axis=1)
        gt_landmarks_cpu = gt_landmarks.cpu().numpy()
        gt_tensor_cpu = gt_tensor.cpu().numpy()
        pred_tensor_cpu = pred_tensor.detach().cpu().numpy()

        del input_tensor, gt_tensor, gt_landmarks, pred_tensor

        for num in range(cur_bath_size):
            gt_landmark = gt_landmarks_cpu[num]
            pred_slices = pred_tensor_cpu[num]
            landmark_num = pred_slices.shape[0]
            number_slice = number_list[num]
            ## for second stage
            pd_slice_list.append(number_slice)
            pd1_hmap_list.append(pred_slices[0])
            pd2_hmap_list.append(pred_slices[1])
            gt_hmap_list.append(gt_tensor_cpu[num])
            gt_landmark_list.append(gt_landmarks_cpu[num])
            # calculate score
            for i in range(landmark_num):
                pred_landmark = data.convert_heatmap_to_landmark(pred_slices[i, :, :])  * int(input_tensor_cpu[num][0].shape[0] / pred_slices.shape[1]) ## 128 -> 512 size
                dist = np.linalg.norm(gt_landmark-pred_landmark)
                dist_epoch += dist
        
        if idx_data == cur_bath_size or idx_data == (cur_bath_size*2):
            title = f'/{int(cur_bath_size/idx_data)}/Epoch_{epoch}'
            os.makedirs(f'{img_save_dir}/{int(cur_bath_size/idx_data)}/', exist_ok=True)
            path_save = img_save_dir + title

            data.save_result_image(path_save, input_tensor_cpu, gt_landmarks_cpu, pred_tensor_cpu, number_list)
        torch.cuda.empty_cache()
    
    loss_epoch /= idx_batch
    dist_epoch /= idx_data * landmark_num

    input_tensor_sec_b_1 = []
    test_list = []
    input_tensor_sec_b_2 = []
    gt_tensor_sec_b_1 = []
    gt_tensor_sec_b_2 = []
    gt_landmark_sec_b_1 = []
    gt_landmark_sec_b_2 = []
    data_num = len(pd_slice_list) 
    number_list = []
    loss_epoch_2 = 0
    dist_epoch_2 = 0
    idx_data = 0
    for pd in pd_slice_list:
        idx_data += 1
        number_list.append(pd)
        pd_year = pd.split('_')[0]
        pd_patient = pd.split('_')[1]
        pd_slice = pd.split('_')[2]
        pd_slice_pre = int(pd_slice) - 1
        str_num = len(str(pd_slice_pre))
        zero_need = 4 - str_num
        z = ''
        for i in range(zero_need):
            z += '0'
        pd_slice_pre = f"{z}{pd_slice_pre}"

        pd_slice_post = int(pd_slice) + 1
        str_num = len(str(pd_slice_post))
        zero_need = 4 - str_num
        z = ''
        for i in range(zero_need):
            z += '0'
        pd_slice_post = f"{z}{pd_slice_post}"
        pd_pre = f"{pd_year}_{pd_patient}_{pd_slice_pre}"
        pd_post = f"{pd_year}_{pd_patient}_{pd_slice_post}"
        input_second_1 = []
        input_second_2 = []
        gt_second_1 = []
        gt_second_2 = []
        gt_second_landm_1 = []
        gt_second_landm_2 = []
        if pd_pre in pd_slice_list:
            input_second_1.append(cv2.resize(pd1_hmap_list[pd_slice_list.index(pd_pre)], (512,512), interpolation=cv2.INTER_CUBIC))
            input_second_2.append(cv2.resize(pd2_hmap_list[pd_slice_list.index(pd_pre)], (512,512), interpolation=cv2.INTER_CUBIC))
        elif pd_pre + '_ZERO' in pd_slice_list:
            input_second_1.append(cv2.resize(pd1_hmap_list[pd_slice_list.index(pd_pre + '_ZERO')], (512,512), interpolation=cv2.INTER_CUBIC))
            input_second_2.append(cv2.resize(pd2_hmap_list[pd_slice_list.index(pd_pre + '_ZERO')], (512,512), interpolation=cv2.INTER_CUBIC))
        else:
            input_second_1.append(cv2.resize(pd1_hmap_list[pd_slice_list.index(pd)], (512,512), interpolation=cv2.INTER_CUBIC))
            input_second_2.append(cv2.resize(pd2_hmap_list[pd_slice_list.index(pd)], (512,512), interpolation=cv2.INTER_CUBIC))
        input_second_1.append(cv2.resize(pd1_hmap_list[pd_slice_list.index(pd)], (512,512), interpolation=cv2.INTER_CUBIC))
        input_second_2.append(cv2.resize(pd2_hmap_list[pd_slice_list.index(pd)], (512,512), interpolation=cv2.INTER_CUBIC))
        # gt_second_1.append(gt_hmap_list[pd_slice_list.index(pd)][0])
        # gt_second_2.append(gt_hmap_list[pd_slice_list.index(pd)][1])
        # gt_second_landm_1.append(gt_landmark_list[pd_slice_list.index(pd)][0])
        # gt_second_landm_2.append(gt_landmark_list[pd_slice_list.index(pd)][1])
        if pd_post in pd_slice_list:
            input_second_1.append(cv2.resize(pd1_hmap_list[pd_slice_list.index(pd_post)], (512,512), interpolation=cv2.INTER_CUBIC))
            input_second_2.append(cv2.resize(pd2_hmap_list[pd_slice_list.index(pd_post)], (512,512), interpolation=cv2.INTER_CUBIC))
        elif pd_post + '_ZERO' in pd_slice_list:
            input_second_1.append(cv2.resize(pd1_hmap_list[pd_slice_list.index(pd_post + '_ZERO')], (512,512), interpolation=cv2.INTER_CUBIC))
            input_second_2.append(cv2.resize(pd2_hmap_list[pd_slice_list.index(pd_post + '_ZERO')], (512,512), interpolation=cv2.INTER_CUBIC))
        else:
            input_second_1.append(cv2.resize(pd1_hmap_list[pd_slice_list.index(pd)], (512,512), interpolation=cv2.INTER_CUBIC))
            input_second_2.append(cv2.resize(pd2_hmap_list[pd_slice_list.index(pd)], (512,512), interpolation=cv2.INTER_CUBIC))
        input_tensor_sec_b_1.append(input_second_1)
        input_tensor_sec_b_2.append(input_second_2)
        # test_list.append([input_second_1])
        gt_tensor_sec_b_1.append(gt_hmap_list[pd_slice_list.index(pd)][0])
        gt_tensor_sec_b_2.append(gt_hmap_list[pd_slice_list.index(pd)][1])
        gt_landmark_sec_b_1.append(gt_landmark_list[pd_slice_list.index(pd)][0])
        gt_landmark_sec_b_2.append(gt_landmark_list[pd_slice_list.index(pd)][1])

        if len(input_tensor_sec_b_1) == batch_size:
            input_tensor_sec_1 = torch.tensor(input_tensor_sec_b_1).to(device)
            input_tensor_sec_2 = torch.tensor(input_tensor_sec_b_2).to(device)
            # test_2 = torch.tensor(test_list).to(device).squeeze()
            # print(input_tensor_sec_1.shape)
            # print(test_2.shape)
            # exit()
            if opti is not None:
                opti.zero_grad()
                pred_tensor = network_2(input_tensor_sec_1)   # torch.Size([batch, 2, 128, 128])
                gt_tensor = torch.tensor(gt_tensor_sec_b_1).to(device) # torch.Size([batch, 128, 128])
                loss_1 = loss_func(pred_tensor[:, 0, :, :], gt_tensor)
                if loss_weight is not None:
                    for number in number_list:
                        if "ZERO" not in number:
                            loss_1 *= float(loss_weight)
                
                loss_1.backward()
                opti.step()

                # calculate score
                pred_tensor_cpu_1 = pred_tensor[:, 0, :, :].detach().cpu().numpy()
                for num in range(batch_size):
                    gt_landmark = gt_landmark_sec_b_1[num]
                    pred_slices = pred_tensor_cpu_1[num]
                    pred_landmark = data.convert_heatmap_to_landmark(pred_slices)  * int(input_tensor_sec_1[num][0].shape[0] / pred_slices.shape[1]) ## 128 -> 512 size
                    dist = np.linalg.norm(gt_landmark-pred_landmark)
                    dist_epoch_2 += dist

                opti.zero_grad()
                pred_tensor = network_2(input_tensor_sec_2)   # torch.Size([batch, 2, 128, 128])
                gt_tensor = torch.tensor(gt_tensor_sec_b_2).to(device).squeeze()
                loss_2 = loss_func(pred_tensor[:, 1, :, :], gt_tensor)
                if loss_weight is not None:
                    for number in number_list:
                        if "ZERO" not in number:
                            loss_2 *= float(loss_weight)
                loss_2.backward()
                opti.step()
                
                # calculate score
                pred_tensor_cpu_2 = pred_tensor[:, 1, :, :].detach().cpu().numpy()
                for num in range(batch_size):
                    gt_landmark = gt_landmark_sec_b_2[num]
                    pred_slices = pred_tensor_cpu_2[num]
                    pred_landmark = data.convert_heatmap_to_landmark(pred_slices)  * int(input_tensor_sec_1[num][0].shape[0] / pred_slices.shape[1]) ## 128 -> 512 size
                    dist = np.linalg.norm(gt_landmark-pred_landmark)
                    dist_epoch_2 += dist

                loss = (loss_1 + loss_2) / 2

                loss_epoch_2 += loss.item()
            # test mode
            else:
                pred_tensor = network_2(input_tensor_sec_1)   # torch.Size([batch, 2, 128, 128])
                gt_tensor = torch.tensor(gt_tensor_sec_b_1).to(device)
                loss_1 = loss_func(pred_tensor[:, 0, :, :], gt_tensor)
                if loss_weight is not None:
                    for number in number_list:
                        if "ZERO" not in number:
                            loss_1 *= float(loss_weight)
                
                # calculate score
                pred_tensor_cpu_1 = pred_tensor[:, 0, :, :].detach().cpu().numpy()
                for num in range(batch_size):
                    gt_landmark = gt_landmark_sec_b_1[num]
                    pred_slices = pred_tensor_cpu_1[num]
                    pred_landmark = data.convert_heatmap_to_landmark(pred_slices)  * int(input_tensor_sec_1[num][0].shape[0] / pred_slices.shape[1]) ## 128 -> 512 size
                    dist = np.linalg.norm(gt_landmark-pred_landmark)
                    dist_epoch_2 += dist

                pred_tensor = network_2(input_tensor_sec_2)   # torch.Size([batch, 2, 128, 128])
                gt_tensor = torch.tensor(gt_tensor_sec_b_2).to(device)
                loss_2 = loss_func(pred_tensor[:, 1, :, :], gt_tensor)
                if loss_weight is not None:
                    for number in number_list:
                        if "ZERO" not in number:
                            loss_2 *= float(loss_weight)
                # calculate score
                pred_tensor_cpu_2 = pred_tensor[:, 1, :, :].detach().cpu().numpy()
                for num in range(batch_size):
                    gt_landmark = gt_landmark_sec_b_2[num]
                    pred_slices = pred_tensor_cpu_2[num]
                    pred_landmark = data.convert_heatmap_to_landmark(pred_slices)  * int(input_tensor_sec_2[num][0].shape[0] / pred_slices.shape[1]) ## 128 -> 512 size
                    dist = np.linalg.norm(gt_landmark-pred_landmark)
                    dist_epoch_2 += dist

                loss = (loss_1 + loss_2) / 2

                loss_epoch_2 += loss.item()
            ## save img    
            if idx_data == batch_size or idx_data == (batch_size*2):
                # pred_tensor_cpu = np.append(pred_tensor_cpu_1, [pred_tensor_cpu_2], axis=1)
                # print("")
                # print(input_tensor_sec_1.shape)
                tmp_tensor_cpu = torch.tensor(np.zeros((batch_size,1,128,128))).to(device).cpu().numpy()
                tmp_gt_cpu = torch.tensor(np.zeros((batch_size,2,1))).to(device).cpu().numpy()
                input_tensor_cpu = input_tensor_sec_1.cpu().numpy()[:, 1, :, :]
                input_tensor_cpu = np.expand_dims(input_tensor_cpu, axis=1)
                # print(input_tensor_cpu.shape)
                pred_tensor_cpu_1 = np.expand_dims(pred_tensor_cpu_1, axis=1)
                # print(tmp_tensor_cpu.shape)
                # print(pred_tensor_cpu_1.shape)
                pred_tensor_cpu = np.append(pred_tensor_cpu_1, tmp_tensor_cpu, axis=1)
                # print(pred_tensor_cpu.shape)
                gt_landmark_1 = torch.tensor(gt_landmark_sec_b_1).to(device).cpu().numpy()
                # print(gt_landmark_1.shape)
                gt_landmark_1 = np.expand_dims(gt_landmark_1, axis=2)
                # print(gt_landmark_1.shape)
                # print(tmp_gt_cpu.shape)
                gt_landmark = np.append(gt_landmark_1, tmp_gt_cpu, axis=2)
                # print("GT =", gt_landmark)
                # print("GT =", gt_landmark.shape)

                os.makedirs(f'{img_save_dir}/{int(batch_size/idx_data)}/', exist_ok=True)
                title = f'/{int(batch_size/idx_data)}/Epoch_{epoch}_Second_1'
                path_save = img_save_dir + title
                data.save_result_image(path_save, input_tensor_cpu, gt_landmark, pred_tensor_cpu, number_list)
                ### 2
                input_tensor_cpu = input_tensor_sec_2.cpu().numpy()[:, 1, :, :]
                input_tensor_cpu = np.expand_dims(input_tensor_cpu, axis=1)
                pred_tensor_cpu_2 = np.expand_dims(pred_tensor_cpu_2, axis=1)
                pred_tensor_cpu = np.append(pred_tensor_cpu_2, tmp_tensor_cpu, axis=1)
                gt_landmark_2 = torch.tensor(gt_landmark_sec_b_2).to(device).cpu().numpy()
                gt_landmark_2 = np.expand_dims(gt_landmark_2, axis=2)
                gt_landmark = np.append(gt_landmark_1, tmp_gt_cpu, axis=2)
                title = f'/{int(batch_size/idx_data)}/Epoch_{epoch}_Second_2'
                path_save = img_save_dir + title
                data.save_result_image(path_save, input_tensor_cpu, gt_landmark, pred_tensor_cpu, number_list)
            
            del input_tensor_sec_1, input_tensor_sec_2, gt_tensor, pred_tensor
            print(f'Progress Sencond Stage: idx_data={idx_data}/{len(data_loader.dataset)}, curr loss={round(loss.item(),4)}', end='\r')
            input_tensor_sec_b_1 = []
            input_tensor_sec_b_2 = []
            gt_landmark_sec_b_1 = []
            gt_landmark_sec_b_2 = []
            gt_tensor_sec_b_1 = []
            gt_tensor_sec_b_2 = []
            number_list = []

            data_num -= batch_size
            if data_num < batch_size:
                batch_size = data_num
        # print(f'Progress Sencond Stage: idx_data={idx_data}/{len(data_loader.dataset)}, curr loss={round(loss.item(),4)}', end='\r')
    # print("")
    # print(f"Data Size = {len(data_loader.dataset)} IDX_Batch = {idx_batch} iDX_Data = {idx_data} landnum = {landmark_num}")
    
    loss_epoch_2 /= idx_batch
    dist_epoch_2 /= (idx_data * landmark_num)
    # print(loss_epoch, dist_epoch, idx_step, loss_epoch_2, dist_epoch_2)
    return [loss_epoch, dist_epoch, idx_step, loss_epoch_2, dist_epoch_2]

def learning_add_classify(data_loader, network, device, opti, loss_func, loss_func_c, loss_weight_c, summary, img_save_dir, epoch, loss_weight, idx_step=0):
    loss_epoch = 0
    loss_hm_epoch = 0
    loss_cf_epoch = 0
    dist_epoch = 0
    idx_batch = 0
    idx_data = 0
    tumor_correct = 0
    zero_correct = 0
    tumor_num = 0
    zero_num = 0
    for batch in data_loader:
        input_tensor = batch['ct'].to(device) # torch.Size([batch, 1, 512, 512])
        gt_tensor = batch['hms'].to(device) # torch.Size([batch, 2, 128, 128])
        gt_landmarks = batch['landmarks'].to(device) 
        number_list = batch['number']
        cur_bath_size = input_tensor.shape[0]

        input_tensor_cpu = input_tensor.cpu().numpy()
        gt_landmarks_cpu = gt_landmarks.cpu().numpy()

        gt_list = []
        for num in range(len(number_list)):
            if 'ZERO' in number_list[num]:
                gt_list.append(0)
                zero_num += 1 
            else:
                gt_list.append(1)
   
        # training mode
        if opti is not None:
            opti.zero_grad()
            pred_tensor = network(input_tensor)  
            pred_tensor_cpu = pred_tensor.detach().cpu().numpy()
            ## detecting hm
            loss_hm = loss_func(pred_tensor, gt_tensor)
            if loss_weight is not None:
                for number in number_list:
                    if "ZERO" not in number:
                        loss_hm *= float(loss_weight)
            
            ## classification
            classifi_input = np.zeros((cur_bath_size))
            for num in range(cur_bath_size):
                pred_slices = pred_tensor_cpu[num]
                landmark_num = pred_slices.shape[0]
                pred_landmark = []
                for i in range(landmark_num):
                    pred_landmark.append(data.convert_heatmap_to_landmark(pred_slices[i, :, :])  * int(input_tensor_cpu[num][0].shape[0] / pred_slices.shape[1])) ## 128 -> 512 size
                
                pd_landmarks = np.array(pred_landmark, dtype=int)
                for a in range(2):
                    for b in range(2):
                        if pd_landmarks[a][b] > 20:
                            classifi_input[num] = 1
                            break
           
            gt_tensor_c = torch.tensor(gt_list).float().to(device= device) # torch.Size([batch])
            classfy_tensor = torch.tensor(classifi_input).float().to(device= device)
            loss_c = loss_func_c(classfy_tensor, gt_tensor_c) * loss_weight_c
            
            loss = loss_hm + loss_c 
            loss.backward()
            opti.step()
        # test mode
        else:
            pred_tensor = network(input_tensor)  
            pred_tensor_cpu = pred_tensor.detach().cpu().numpy()

            ## predict hm
            loss_hm = loss_func(pred_tensor, gt_tensor)
            if loss_weight is not None:
                for number in number_list:
                    if "ZERO" not in number:
                        loss_hm *= float(loss_weight)
            
            ## classification
            classifi_input = np.zeros((cur_bath_size))
            for num in range(cur_bath_size):
                pred_slices = pred_tensor_cpu[num]
                landmark_num = pred_slices.shape[0]
                pred_landmark = []
                for i in range(landmark_num):
                    pred_landmark.append(data.convert_heatmap_to_landmark(pred_slices[i, :, :])  * int(input_tensor_cpu[num][0].shape[0] / pred_slices.shape[1])) ## 128 -> 512 size
                
                pd_landmarks = np.array(pred_landmark, dtype=int)
                for a in range(2):
                    for b in range(2):
                        if pd_landmarks[a][b] > 20:
                            classifi_input[num] = 1
                            break

            gt_tensor_c = torch.tensor(gt_list).float().to(device= device) # torch.Size([batch])
            classfy_tensor = torch.tensor(classifi_input).float().to(device= device)
            loss_c = loss_func_c(classfy_tensor, gt_tensor_c) * loss_weight_c
            
            loss = loss_hm + loss_c 
        ## calculate classify score
        for idx in range(len(gt_list)):
            if gt_list[idx] == classifi_input[idx]:    
                if gt_list[idx] == 0:
                    zero_correct += 1
                else:
                    tumor_correct += 1
        idx_batch += 1
        idx_step += 1
        idx_data += input_tensor.shape[0]

        loss_epoch += loss.item()
        loss_hm_epoch += loss_hm.item()
        loss_cf_epoch += loss_c.item()
        if opti is not None:
            summary.add_scalar('Loss/Step_train', loss.item(), idx_step)
        else:
            summary.add_scalar('Loss/Step_test', loss.item(), idx_step)
        print(f'Progress: idx_data={idx_data}/{len(data_loader.dataset)}, curr loss={round(loss.item(),4)}', end='\r')

        del input_tensor, gt_tensor, gt_landmarks, pred_tensor, gt_tensor_c, classfy_tensor

        for num in range(cur_bath_size):
            gt_landmark = gt_landmarks_cpu[num]
            pred_slices = pred_tensor_cpu[num]
            landmark_num = pred_slices.shape[0]

            # calculate score
            for i in range(landmark_num):
                pred_landmark = data.convert_heatmap_to_landmark(pred_slices[i, :, :])  * int(input_tensor_cpu[num][0].shape[0] / pred_slices.shape[1]) ## 128 -> 512 size
                dist = np.linalg.norm(gt_landmark-pred_landmark)
                dist_epoch += dist
        
        if idx_data == cur_bath_size or idx_data == (cur_bath_size*2):
            title = f'/{int(cur_bath_size/idx_data)}/Epoch_{epoch}'
            os.makedirs(f'{img_save_dir}/{int(cur_bath_size/idx_data)}/', exist_ok=True)
            path_save = img_save_dir + title 
            data.save_result_image_classification(path_save, input_tensor_cpu, gt_landmarks_cpu, pred_tensor_cpu, number_list, classifi_input)

        torch.cuda.empty_cache()
    
    loss_epoch /= idx_batch
    loss_hm_epoch /= idx_batch
    loss_cf_epoch /= idx_batch
    dist_epoch /= idx_data * landmark_num
    tumor_num = len(data_loader.dataset) - zero_num
    classify_acc = (tumor_correct + zero_correct) / len(data_loader.dataset)
    tumor_correct /= tumor_num
    zero_correct /= zero_num


    return [loss_epoch, loss_hm_epoch, loss_cf_epoch, dist_epoch, idx_step, classify_acc, tumor_correct, zero_correct]

if __name__ == "__main__":
    # train()
    # train_3slice_1stg()
    train_add_classifier()